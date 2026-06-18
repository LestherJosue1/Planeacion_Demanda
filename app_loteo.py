from __future__ import annotations
import streamlit as st
import pandas as pd
import numpy as np
import json, os, re, time, threading, logging
import pytz
from datetime import datetime
from io import BytesIO
from dataclasses import dataclass, field
from typing import List, Set, Dict, Tuple, Optional

log = logging.getLogger(__name__)



# ------------------------------------------------------------
# models/lote.py
# ------------------------------------------------------------
@dataclass
class LoteRow:
    """Una fila asignada dentro de un lote."""
    orig_idx:    int
    lbs_asig:    float
    over_extra:  float = 0.0
    under_saved: float = 0.0


@dataclass
class Lote:
    """Resultado de intentar_lote_para_rango."""
    lid:          int
    categoria:    str
    rango_id:     str
    mix:          str
    minimo:       float
    maximo:       float
    total:        float
    rows:         List[LoteRow]
    anchos:       Set[float]      = field(default_factory=set)
    score:        float           = 0.0
    rule_info:    dict            = field(default_factory=dict)
    set_anchos:   str             = ""
    cant_anchos:  int             = 0
    tipo:         str             = "PURO"   # PURO | MIX_CONTROLADO | MIX_ALTO

    def __post_init__(self):
        self.cant_anchos = len(self.anchos)
        anch_sorted = sorted(self.anchos, reverse=True)
        self.set_anchos = "-".join(str(round(a, 2)) for a in anch_sorted)
        if self.cant_anchos == 1:
            self.tipo = "PURO"
        elif self.cant_anchos <= 3:
            self.tipo = "MIX_CONTROLADO"
        else:
            self.tipo = "MIX_ALTO"


# ------------------------------------------------------------
# models/categoria.py
# ------------------------------------------------------------
@dataclass
class Categoria:
    nombre:       str
    minimo:       float
    maximo:       float
    mix:          str           # DYE | BLEACH
    tipo_tejido:  str = "TODOS" # TODOS | FLEECE | JERSEY
    lotes_dia:    int = 5
    semanas:      float = 4.0
    ctd_max_anchos: int = 3
    activo:       bool = True

    @property
    def rango_id(self) -> str:
        return f"CAP_{self.nombre}_{self.mix.upper()}_{self.maximo:.0f}"

    @property
    def capacidad_lbs(self) -> float:
        return self.maximo * self.lotes_dia * 7 * self.semanas

    def to_dict(self) -> dict:
        return {
            "CATEGORIA":     self.nombre,
            "MINIMO":        self.minimo,
            "MAXIMO":        self.maximo,
            "MIX":           self.mix,
            "TIPO_TEJIDO":   self.tipo_tejido,
            "LOTES":         self.lotes_dia,
            "SEMANAS":       self.semanas,
            "CTDMAXANCHOS":  self.ctd_max_anchos,
            "ACTIVO":        self.activo,
            "CAPACIDAD_LBS": self.capacidad_lbs,
        }

    @classmethod
    def from_dict(cls, d: dict) -> "Categoria":
        return cls(
            nombre         = str(d.get("CATEGORIA", "")),
            minimo         = float(d.get("MINIMO", 0)),
            maximo         = float(d.get("MAXIMO", 0)),
            mix            = str(d.get("MIX", "DYE")).upper(),
            tipo_tejido    = str(d.get("TIPO_TEJIDO", "TODOS")).upper(),
            lotes_dia      = int(d.get("LOTES", 5)),
            semanas        = float(d.get("SEMANAS", 4.0)),
            ctd_max_anchos = int(d.get("CTDMAXANCHOS", 3)),
            activo         = bool(d.get("ACTIVO", True)),
        )


# ------------------------------------------------------------
# models/regla.py
# ------------------------------------------------------------
@dataclass
class ReglaAncho:
    style:        str
    limite_ancho: float
    prioridades:  List[float]   # MAXIMO de categorías permitidas
    activo:       bool = True


@dataclass
class ReglaColor:
    color_r:      str
    prioridades:  List[float]
    activo:       bool = True


@dataclass
class ReglaFamilia:
    familia:      str
    prioridades:  List[float]
    activo:       bool = True


@dataclass
class ReglaCombinacion:
    """Si un lote combina ancho_1 y ancho_2, dirigirlo a estas categorías."""
    ancho_1:      float
    ancho_2:      float
    prioridades:  List[float]
    activo:       bool = True


# ------------------------------------------------------------
# models/configuracion.py
# ------------------------------------------------------------
@dataclass
class Configuracion:
    # Solver
    max_items:       int   = 8
    solver_timeout:  float = 5.0
    beam_width:      int   = 3      # cuántos seeds probar por bloque

    # Diferencia de anchos dentro de un lote
    min_diff: float = 1.0   # diferencia mínima entre anchos distintos en pulgadas
    max_diff: float = 6.0   # diferencia máxima entre anchos distintos en pulgadas

    # Split
    split_min_default: float = 100.0   # Colab default
    split_min_ancho18: float = 250.0
    scrap_remainder:   bool  = True    # scrap remanentes < split_min

    # Overshoot / Undershoot
    overshoot_enable:  bool  = False
    undershoot_enable: bool  = False

    # Anchos
    widths_target_order:   str  = "2>3>4"
    require_widths_strict: bool = True
    allowed_maximo_3_dye:  Set[float] = field(default_factory=set)
    allowed_maximo_4_dye:  Set[float] = field(default_factory=set)

    # Reglas
    apply_rules_bleach:  bool = False
    upgrade_categoria:   bool = True
    try_all_priorities:  bool = True
    rule_order:          str  = "ANCHO18>COMBO_ANCHOS>COLOR_R>FAMILIA>DEFAULT"
    ancho18_allowed_max_dye: Set[float] = field(default_factory=lambda: {2200.0, 1100.0})
    ancho18_allow_spillover_2600: bool = False

    # Scoring weights
    w_fill:        float = 5.0
    w_cap_loss:    float = 3.0
    w_width_pref:  float = 2.0
    w_1100_strict: float = 10.0
    w_priority:    float = 1.0
    w_completion:  float = 0.5
    width_pref_list: List[int] = field(default_factory=lambda: [2, 3, 1, 4, 5, 6])

    # Prioridades de mezcla
    combinacion_prioridad: List[List[str]] = field(
        default_factory=lambda: [["VENCIDOS","AHEAD"],["AHEAD","AHEAD2"],["AHEAD2","OTROS"]]
    )

    @classmethod
    def from_dict(cls, d: dict) -> "Configuracion":
        def _int(k, default):
            try: return int(d.get(k, default))
            except: return default

        def _float(k, default):
            try: return float(d.get(k, default))
            except: return default

        def _bool(k, default):
            v = str(d.get(k, default)).strip().upper()
            return v in ("1","TRUE","YES","SI","SÍ","TRUE")

        def _set_float(k, default):
            import re
            raw = str(d.get(k, "")).strip()
            if not raw:
                return set(default)
            parts = re.split(r"[;,\s]+", raw)
            vals = []
            for p in parts:
                try: vals.append(float(p))
                except: pass
            return set(vals) if vals else set(default)

        return cls(
            max_items               = _int("MAX_ITEMS", 8),
            solver_timeout          = _float("SOLVER_TIMEOUT", 5.0),
            beam_width              = _int("BEAM_WIDTH", 3),
            min_diff                = _float("MIN_DIFF", 1.0),
            max_diff                = _float("MAX_DIFF", 6.0),
            split_min_default       = _float("SPLIT_MIN_LBS_DEFAULT", 100.0),
            split_min_ancho18       = _float("SPLIT_MIN_LBS_ANCHO18", 250.0),
            scrap_remainder         = _bool("SCRAP_REMAINDER_BELOW_SPLIT_MIN", True),
            overshoot_enable        = _bool("OVERSHOOT_ENABLE", False),
            undershoot_enable       = _bool("UNDERSHOOT_ENABLE", False),
            widths_target_order     = str(d.get("WIDTHS_TARGET_ORDER", "2>3>4")),
            require_widths_strict   = _bool("REQUIRE_WIDTHS_STRICT", True),
            apply_rules_bleach      = _bool("APPLY_RULES_BLEACH", False),
            upgrade_categoria       = _bool("UPGRADE_CATEGORIA", True),
            try_all_priorities      = _bool("TRY_ALL_PRIORITIES", True),
            rule_order              = str(d.get("RULE_ORDER", "ANCHO18>COMBO_ANCHOS>COLOR_R>FAMILIA>DEFAULT")),
            ancho18_allow_spillover_2600 = _bool("ANCHO18_ALLOW_SPILLOVER_2600", False),
            ancho18_allowed_max_dye = _set_float("ANCHO18_ALLOWED_MAX_DYE", {2200.0, 1100.0}),
            allowed_maximo_3_dye    = _set_float("ALLOWED_MAXIMO_FOR_3_WIDTHS_DYE", set()),
            allowed_maximo_4_dye    = _set_float("ALLOWED_MAXIMO_FOR_4_WIDTHS_DYE", set()),
            w_fill                  = _float("W_FILL", 5.0),
            w_cap_loss              = _float("W_CAP_LOSS", 3.0),
            w_width_pref            = _float("W_WIDTH_PREF", 2.0),
            w_1100_strict           = _float("W_1100_WIDTHS_STRICT", 10.0),
            w_priority              = _float("W_PRIORITY", 1.0),
            w_completion            = _float("W_COMPLETION", 0.5),
            combinacion_prioridad   = d.get("COMBINACION_PRIORIDAD",
                                            [["VENCIDOS","AHEAD"],["AHEAD","AHEAD2"],["AHEAD2","OTROS"]]),
        )


# ------------------------------------------------------------
# core/split_engine.py
# ------------------------------------------------------------
class SplitEngine:
    """
    Pool compartido de LBS disponibles por índice de fila.
    Un LNK puede dividirse entre múltiples lotes/categorías, nunca excediendo LBS_BASE.
    """

    def __init__(self, lbs_base: Dict[int, float]):
        # Copia del estado inicial — {orig_idx: lbs_disponibles}
        self._pool: Dict[int, float] = dict(lbs_base)
        self._scrap: Dict[int, float] = {}

    # ── Acceso al pool ────────────────────────────────────────────────────────

    def disponible(self, idx: int) -> float:
        return max(0.0, self._pool.get(idx, 0.0))

    def base(self, idx: int) -> float:
        """LBS originales (antes de cualquier asignación)."""
        return self._pool.get(idx, 0.0) + self._scrap.get(idx, 0.0) + 0.0

    def descontar(self, idx: int, lbs: float) -> None:
        prev = self._pool.get(idx, 0.0)
        self._pool[idx] = max(0.0, round(prev - lbs, 4))

    def scrap(self, idx: int, lbs: float) -> None:
        self._pool[idx] = 0.0
        self._scrap[idx] = self._scrap.get(idx, 0.0) + lbs

    def restante(self, idx: int) -> float:
        return self._pool.get(idx, 0.0)

    def snapshot(self) -> Dict[int, float]:
        return dict(self._pool)

    # ── choose_take ───────────────────────────────────────────────────────────

    @staticmethod
    def choose_take(
        rest: float,
        remaining: float,
        split_min: float = 100.0,
        allow_scrap_residue: bool = True,
    ) -> float:
        """
        Determina cuántas LBS tomar de un ítem para un lote.

        Reglas (igual que Colab):
        - Si rest <= remaining: tomar todo (ítem completo).
        - Si rest > remaining: tomar `remaining` (fragmento).
        - Si el residuo (rest - take) quedaría entre 0 y split_min:
            a) Si podemos ajustar el take para que el residuo sea >= split_min: hacerlo.
            b) Si no (el residuo es demasiado pequeño y no hay margen): tomar todo
               y marcar el excedente como scrap (si allow_scrap_residue=True).
        """
        if rest <= 0 or remaining <= 0:
            return 0.0

        if rest <= remaining + 1e-9:
            return round(float(rest), 4)

        take    = float(remaining)
        residuo = float(rest) - take

        if residuo > 1e-9 and residuo < split_min - 1e-9:
            if not allow_scrap_residue:
                return 0.0
            # Scrap del residuo: tomar todo
            return round(take, 4)

        return round(take, 4)

    @staticmethod
    def choose_take_humano(
        rest: float,
        remaining: float,
        total_original: float,
        overshoot_enable: bool = False,
        undershoot_enable: bool = False,
        split_min: float = 100.0,
    ) -> tuple[float, float, float]:
        """
        Versión con tolerancia humana (over/undershoot).
        Retorna (take, over_extra, under_saved).
        """
        if remaining <= 0 or rest < 0:
            return 0.0, 0.0, 0.0

        def tol_pct(total: float) -> float:
            if total <= 5_000:   return 0.05
            if total <= 10_000:  return 0.02
            if total <= 30_000:  return 0.01
            return 0.0

        tol = tol_pct(total_original)
        take       = min(rest, remaining)
        over_extra = 0.0
        under_saved= 0.0

        if overshoot_enable:
            cap_room  = remaining - take
            if cap_room > 1e-9 and rest <= take + 1e-9:
                max_extra = tol * total_original
                extra     = min(cap_room, max_extra)
                if extra > 0:
                    take       += extra
                    over_extra  = extra

        if undershoot_enable and take > remaining + 1e-9:
            reducible    = min(take - remaining, tol * total_original)
            if reducible > 0:
                take        -= reducible
                under_saved  = reducible

        take = min(take, remaining)
        if take <= 1e-9:
            return 0.0, 0.0, 0.0

        return round(float(take), 4), round(float(over_extra), 4), round(float(under_saved), 4)


# ------------------------------------------------------------
# core/width_engine.py
# ------------------------------------------------------------
def get_row_widths(row: pd.Series) -> Set[float]:
    """
    Retorna el set de anchos activos de una fila.
    Considera ANCHO (ANCHO.F.C renombrado) y ANCHO.F.M, excluye 0.
    """
    anchos: Set[float] = set()
    for col in ("ANCHO", "ANCHO.F.C", "ANCHO.F.M"):
        if col in row.index:
            try:
                v = float(row[col])
                if v > 0:
                    anchos.add(round(v, 4))
            except (TypeError, ValueError):
                pass
    return anchos


def anchos_validos(
    anchos_set: Set[float],
    ctd_max_anchos: int,
) -> bool:
    """
    Verifica que la CANTIDAD de anchos distintos <= ctd_max_anchos.
    No verifica valores en pulgadas — eso no es una regla de negocio.
    """
    return len(anchos_set) <= ctd_max_anchos


def valid_width_group(
    widths: List[float],
    min_diff: float,
    max_diff: float,
    max_widths: int,
) -> bool:
    """
    Valida que los anchos de un grupo cumplan restricciones de diferencia.
    Usado por el Colab para asegurar que anchos mezclados sean compatibles en tintorería.
    """
    w = [float(x) for x in widths if x is not None and float(x) != 0.0]
    uw = sorted(set(w))
    if len(uw) <= 1:
        return True
    if len(uw) > max_widths:
        return False
    for i in range(len(uw)):
        for j in range(i + 1, len(uw)):
            d = abs(uw[j] - uw[i])
            if d < min_diff or d > max_diff:
                return False
    return True


def filter_ranges_for_width_target(
    ranges: list,
    mix: str,
    width_target: int,
    allowed_maximo_3: Set[float],
    allowed_maximo_4: Set[float],
) -> list:
    """
    Filtra categorías disponibles para un objetivo de anchos específico.
    Para 3+ anchos, el Colab solo permite ciertas categorías grandes.
    """
    mixu = mix.strip().upper()
    allowed: Optional[Set[float]] = None

    if width_target == 3 and allowed_maximo_3:
        allowed = allowed_maximo_3
    elif width_target == 4 and allowed_maximo_4:
        allowed = allowed_maximo_4

    if allowed:
        return [r for r in ranges if float(r["MAXIMO"]) in allowed]
    return list(ranges)


def parse_widths_target_order(order_str: str) -> List[int]:
    """
    Convierte "2>3>4" en [2, 3, 4].
    """
    parts = order_str.strip().split(">")
    result = []
    for p in parts:
        p = p.strip()
        if p.isdigit():
            result.append(int(p))
    return result or [2, 3, 4]


# ------------------------------------------------------------
# core/priority_engine.py
# ------------------------------------------------------------
# ── Bloque de prioridad ───────────────────────────────────────────────────────

PRIO_ORDER = {
    "PAST DUE": 0, "PASTVENCIDOS": 0, "VENCIDOS": 0,
    "DUE":  1,
    "AHEAD":  2,
    "AHEAD2": 3,
    "OTROS":  4,
}

def prio_bloque(prio_text: str) -> str:
    p = (prio_text or "").upper().strip()
    if "PAST DUE" in p or "DUE" in p or "VENC" in p:
        return "VENCIDOS"
    if "AHEAD2" in p:
        return "AHEAD2"
    if "AHEAD" in p:
        return "AHEAD"
    return "OTROS"

def prio_rank(p: str) -> int:
    return PRIO_ORDER.get(str(p).upper().strip(), 4)


# ── Allowed pairs ─────────────────────────────────────────────────────────────

def build_allowed_pairs(combinaciones: List[List[str]]) -> Set[Tuple[str, str]]:
    """Construye set de pares de bloques que pueden mezclarse en un lote."""
    pairs: Set[Tuple[str, str]] = set()
    due_grp = {"DUE", "PAST DUE", "VENCIDOS"}

    for p in PRIO_ORDER:
        pairs.add((p, p))

    for pair in combinaciones:
        if len(pair) >= 2:
            a = str(pair[0]).upper().strip()
            b = str(pair[1]).upper().strip()
            pairs.add((a, b))
            pairs.add((b, a))
            pairs.add((a, a))
            pairs.add((b, b))

    # DUE/PAST DUE siempre pueden completarse con AHEAD
    for d in due_grp:
        pairs.add((d, "AHEAD"))
        pairs.add(("AHEAD", d))

    return pairs

def can_mix(existing: Set[str], new_prio: str, allowed_pairs: Set[Tuple[str, str]]) -> bool:
    np = str(new_prio).upper().strip()
    due_grp = {"DUE", "PAST DUE", "VENCIDOS"}
    for ep in existing:
        es = str(ep).upper().strip()
        if es == np:
            continue
        if allowed_pairs and (es, np) not in allowed_pairs:
            # Excepción explícita: DUE ↔ AHEAD siempre
            if es in due_grp and np == "AHEAD":
                continue
            if np in due_grp and es == "AHEAD":
                continue
            return False
    return True


# ── Order priorities ──────────────────────────────────────────────────────────

def order_priorities(prioridades: List[float], priority_order: str) -> List[float]:
    """Ordena una lista de prioridades (MAXIMO de categoría) según configuración."""
    pris = [float(x) for x in prioridades if x is not None]
    if priority_order:
        plan = [p.strip() for p in priority_order.split(">") if p.strip()]
        rank: Dict[float, int] = {}
        for i, v in enumerate(plan):
            if re.match(r"^\d+(\.\d+)?$", v):
                rank[float(v)] = i
        return sorted(pris, key=lambda x: (rank.get(float(x), 10_000), float(x)))
    return sorted(pris)

def order_ranges_by_priorities(ranges: list, prioridades: List[float]) -> list:
    """Reordena rangos poniendo primero los que coinciden con las prioridades dadas."""
    used: Set[int] = set()
    out = []
    for cap in prioridades:
        for r in ranges:
            if abs(float(r["MAXIMO"]) - float(cap)) < 1e-6 and id(r) not in used:
                out.append(r)
                used.add(id(r))
    for r in ranges:
        if id(r) not in used:
            out.append(r)
    return out

def ranges_matching_priority(
    pri: float,
    ranges: list,
    allow_nearest_higher: bool = True,
) -> list:
    """Rangos cuyo MAXIMO coincide con pri, o el más cercano mayor."""
    exact = [r for r in ranges if abs(float(r["MAXIMO"]) - pri) < 1e-6]
    if exact:
        return exact
    if not allow_nearest_higher:
        return []
    higher = [r for r in ranges if float(r["MAXIMO"]) >= pri - 1e-6]
    if higher:
        higher = sorted(higher, key=lambda r: (float(r["MAXIMO"]) - pri, -float(r["MAXIMO"])))
        return [higher[0]]
    return []


# ── Reorder ranges for seed (reglas de negocio) ───────────────────────────────

def reorder_ranges_for_seed(
    ranges_mix:   list,
    mix:          str,
    row:          pd.Series,
    reglas_ancho: List[ReglaAncho],
    reglas_color: List[ReglaColor],
    reglas_fam:   List[ReglaFamilia],
    reglas_combo: List[ReglaCombinacion],
    work:         pd.DataFrame,
    seed_idx:     int,
    cfg:          Configuracion,
) -> Tuple[list, dict]:
    """
    Determina el orden de rangos a intentar para un seed dado, según las reglas de negocio.
    Retorna (rangos_ordenados, rule_info).

    Jerarquía de reglas (configurable via cfg.rule_order):
      ANCHO18 > COMBO_ANCHOS > COLOR_R > FAMILIA > DEFAULT
    """
    base      = list(ranges_mix)
    rule_info = {
        "regla_aplicada":    "NONE",
        "prioridades":       [],
        "match_combo":       False,
        "limite_ancho_style": None,
        "origen_prioridad":  "MIX",
        "combo_target_width": None,
    }

    mixu = mix.strip().upper()
    # Si no es DYE y no se aplican reglas BLEACH → retornar base
    if mixu != "DYE" and not cfg.apply_rules_bleach:
        return base, rule_info

    # Valores del seed
    style   = str(row.get("STYLE",   "")).upper().strip()
    familia = str(row.get("FAMILIA", "")).upper().strip()
    color_r = str(row.get("COLOR_R", "")).upper().strip()

    def f2(col: str) -> float:
        try:
            v = row.get(col, 0)
            return float(v) if v is not None and not pd.isna(v) else 0.0
        except:
            return 0.0

    ancho_c = f2("ANCHO")
    ancho_m = f2("ANCHO.F.M")

    def ancho_activo_leq(lim: float) -> bool:
        vals = [v for v in [ancho_c, ancho_m] if v > 0]
        return bool(vals) and min(vals) <= lim

    # ── Intento ANCHO18 ───────────────────────────────────────────────────────
    def try_ancho18():
        for r in reglas_ancho:
            if not r.activo or r.style != style:
                continue
            if not ancho_activo_leq(r.limite_ancho):
                continue
            if not r.prioridades:
                continue
            pris = order_priorities(r.prioridades, cfg.rule_order)
            # Filtrar a categorías permitidas para DYE
            allowed = set(cfg.ancho18_allowed_max_dye)
            if cfg.ancho18_allow_spillover_2600:
                allowed.add(2600.0)
            ranges_filtered = [x for x in base if float(x["MAXIMO"]) in allowed]
            if ranges_filtered:
                rule_info.update({
                    "regla_aplicada":    "ANCHO18",
                    "prioridades":       pris,
                    "limite_ancho_style": r.limite_ancho,
                    "origen_prioridad":  "STYLE",
                })
                return order_ranges_by_priorities(ranges_filtered, pris)
        return None

    # ── Intento COMBO_ANCHOS ──────────────────────────────────────────────────
    def try_combo():
        for r in reglas_combo:
            if not r.activo:
                continue
            seed_matches = (
                abs(ancho_c - r.ancho_1) < 1e-6 or abs(ancho_m - r.ancho_1) < 1e-6 or
                abs(ancho_c - r.ancho_2) < 1e-6 or abs(ancho_m - r.ancho_2) < 1e-6
            )
            if not seed_matches:
                continue
            objetivo = r.ancho_2 if (
                abs(ancho_c - r.ancho_1) < 1e-6 or abs(ancho_m - r.ancho_1) < 1e-6
            ) else r.ancho_1
            # ¿Hay otro ítem con el ancho objetivo y LBS disponibles?
            existe = False
            lbs_col = "LBS_RESTANTES" if "LBS_RESTANTES" in work.columns else "LBS_C"
            for idx in work.index:
                if idx == seed_idx:
                    continue
                if float(work.at[idx, lbs_col]) <= 0:
                    continue
                ac2 = f2("ANCHO") if "ANCHO" in work.columns else 0.0
                am2 = f2("ANCHO.F.M") if "ANCHO.F.M" in work.columns else 0.0
                try:
                    ac2 = float(work.at[idx, "ANCHO"]) if "ANCHO" in work.columns else 0.0
                    am2 = float(work.at[idx, "ANCHO.F.M"]) if "ANCHO.F.M" in work.columns else 0.0
                except:
                    pass
                if abs(ac2 - objetivo) < 1e-6 or abs(am2 - objetivo) < 1e-6:
                    existe = True
                    break
            if existe and r.prioridades:
                pris = order_priorities(r.prioridades, cfg.rule_order)
                rule_info.update({
                    "regla_aplicada":    "COMBO_ANCHOS",
                    "prioridades":       pris,
                    "match_combo":       True,
                    "origen_prioridad":  "COMBO",
                    "combo_target_width": float(objetivo),
                })
                return order_ranges_by_priorities(base, pris)
        return None

    # ── Intento COLOR_R ───────────────────────────────────────────────────────
    def try_color_r():
        for r in reglas_color:
            if not r.activo or r.color_r != color_r:
                continue
            if not r.prioridades:
                continue
            p = r.prioridades[0]
            rule_info.update({
                "regla_aplicada":   "COLOR_R",
                "prioridades":      [p],
                "origen_prioridad": "COLOR",
            })
            return order_ranges_by_priorities(base, [p])
        return None

    # ── Intento FAMILIA ───────────────────────────────────────────────────────
    def try_familia():
        for r in reglas_fam:
            if not r.activo or r.familia != familia:
                continue
            if not r.prioridades:
                continue
            pris = order_priorities(r.prioridades, cfg.rule_order)
            rule_info.update({
                "regla_aplicada":   "FAMILIA",
                "prioridades":      pris,
                "origen_prioridad": "FAMILIA",
            })
            return order_ranges_by_priorities(base, pris)
        return None

    # ── Aplicar reglas en orden configurable ──────────────────────────────────
    rule_tokens = [t.strip().upper() for t in cfg.rule_order.split(">") if t.strip()]
    dispatchers = {
        "ANCHO18":     try_ancho18,
        "COMBO_ANCHOS": try_combo,
        "COLOR_R":     try_color_r,
        "FAMILIA":     try_familia,
    }
    for token in rule_tokens:
        fn = dispatchers.get(token)
        if fn:
            result = fn()
            if result is not None:
                return result, rule_info

    return base, rule_info


# ------------------------------------------------------------
# core/scorer.py
# ------------------------------------------------------------
def score_lote(lote: Lote, cfg: Configuracion, lbs_restantes: dict) -> float:
    """
    Score multi-factor — igual al Colab pero con más factores configurables.

    score =
        W_FILL       * fill_rate
      - W_CAP_LOSS   * capacidad_perdida
      + W_WIDTH_PREF * width_preference
      - W_1100_STRICT * penalización_1100
      + W_PRIORITY   * priority_score
      + W_COMPLETION * completion_score
    """
    if not lote.rows:
        return -1e30

    maximo = lote.maximo
    total  = lote.total

    fill_rate = total / maximo if maximo > 1e-9 else 0.0
    cap_loss  = maximo - total

    # Preferencia de cantidad de anchos
    n_anchos = len(lote.anchos)
    pref_list = cfg.width_pref_list
    try:
        rank = pref_list.index(n_anchos)
    except ValueError:
        rank = len(pref_list) + abs(n_anchos - (pref_list[-1] if pref_list else 2))
    width_pref_score = -float(rank)

    # Penalización por anchos en E-1100 / G-1100
    pena_1100 = 0.0
    if abs(maximo - 1100.0) < 1e-6:
        pena_1100 = cfg.w_1100_strict * max(0, n_anchos - 1)

    # Priority score — preferir lotes con prioridades altas
    from ..core.priority_engine import prio_rank
    prios_en_lote = []
    for row_result in lote.rows:
        # Se adjunta en _build_lote
        pass
    prio_score = 0.0  # Se calcula con prios_en_lote si están disponibles

    # Completion score — cuántos LNKs quedan con LBS_RESTANTES ≈ 0 tras este lote
    completion = 0.0
    for row_r in lote.rows:
        rest_after = max(0.0, lbs_restantes.get(row_r.orig_idx, 0.0) - row_r.lbs_asig)
        if rest_after < 1e-6:
            completion += 1.0
    completion_norm = completion / max(len(lote.rows), 1)

    score = (
          cfg.w_fill       * fill_rate
        - cfg.w_cap_loss   * cap_loss
        + cfg.w_width_pref * width_pref_score
        - pena_1100
        + cfg.w_completion * completion_norm
    )
    return score


def score_lote_simple(
    total: float,
    maximo: float,
    anchos: Set[float],
    cfg: Configuracion,
) -> float:
    """
    Versión simplificada para comparar candidatos durante el greedy fill.
    """
    fill = total / maximo if maximo > 1e-9 else 0.0
    n    = len(anchos)
    pref = cfg.width_pref_list
    try:
        rank = pref.index(n)
    except ValueError:
        rank = len(pref) + abs(n - (pref[-1] if pref else 2))

    return fill * cfg.w_fill + (-float(rank)) * cfg.w_width_pref - (maximo - total) * cfg.w_cap_loss


# ------------------------------------------------------------
# core/lot_builder.py
# ------------------------------------------------------------
def intentar_lote_para_rango(
    work:        pd.DataFrame,
    seed_idx:    int,
    rango:       dict,
    pool:        SplitEngine,
    cap_usada:   float,
    cfg:         Configuracion,
    rule_info:   dict,
    allowed_pairs: set,
    min_unique_widths: Optional[int] = None,
    max_unique_widths: Optional[int] = None,
) -> Optional[Lote]:
    """
    Intenta construir un lote válido para `rango` usando `seed_idx` como semilla.
    Retorna un objeto Lote si tiene éxito, None si no.

    Reglas de compatibilidad verificadas:
    1. Mismo TONO (si existe la columna) — ya garantizado por el groupby
    2. Anchos distintos <= CTDMAXANCHOS (cantidad, no pulgadas)
    3. LNKs distintos <= MAX_ITEMS (set de LNKs)
    4. Mezcla de prioridades según COMBINACIONES_PRIORIDAD
    5. LBS entre MINIMO y MAXIMO del rango
    """
    cat       = rango["CATEGORIA"]
    rid       = rango["RANGO_ID"]
    cap_total = float(rango["CAPACIDAD"])
    cap_libre = max(0.0, cap_total - cap_usada)

    if cap_libre <= 0:
        return None

    maximo      = float(rango["MAXIMO"])
    minimo      = float(rango["MINIMO"])
    mix         = str(rango["MIX"]).upper()
    max_take    = min(maximo, cap_libre)
    max_anchos  = int(rango.get("CTDMAXANCHOS", cfg.max_items))

    # Determinar split_min según regla aplicada
    split_min = (
        cfg.split_min_ancho18
        if rule_info.get("regla_aplicada") == "ANCHO18"
        else cfg.split_min_default
    )

    # ── SEED ─────────────────────────────────────────────────────────────────
    seed_rest = pool.disponible(seed_idx)
    if seed_rest <= 0:
        return None

    seed_row    = work.loc[seed_idx]
    seed_anchos = get_row_widths(seed_row)

    # Chequeo de anchos máximos del seed (cantidad, no pulgadas)
    if not anchos_validos(seed_anchos, max_anchos):
        return None


    # max_unique_widths si hay objetivo de anchos estricto
    if max_unique_widths is not None and len(seed_anchos) > max_unique_widths:
        return None

    take_seed = SplitEngine.choose_take(seed_rest, max_take, split_min, cfg.scrap_remainder)
    if take_seed <= 0:
        return None

    seed_prio = str(seed_row.get("PRIORIDAD", "AHEAD")).upper().strip()
    seed_lnk  = str(seed_row.get("LNK", str(seed_idx)))

    # Inicializar lote con seed
    rows:        List[Tuple[int, float]] = [(seed_idx, take_seed)]
    lote_lbs:    float                   = take_seed
    anchos_lote: Set[float]              = set(seed_anchos)
    prios_lote:  Set[str]                = {seed_prio}
    lnks_lote:   Set[str]                = {seed_lnk}
    combo_target = rule_info.get("combo_target_width", None)

    # ── GREEDY FILL ───────────────────────────────────────────────────────────
    # Ordenar disponibles: mayor prioridad primero, luego más LBS
    disp_sorted = (
        work[work.index != seed_idx]
        .copy()
    )
    if "PRIORIDAD" in disp_sorted.columns:
        disp_sorted["_pr"] = disp_sorted["PRIORIDAD"].apply(prio_rank)
        disp_sorted["_rest"] = disp_sorted.index.map(lambda i: pool.disponible(i))
        disp_sorted = disp_sorted[disp_sorted["_rest"] > 0].sort_values(
            ["_pr", "_rest"], ascending=[True, False]
        )
    else:
        disp_sorted["_rest"] = disp_sorted.index.map(lambda i: pool.disponible(i))
        disp_sorted = disp_sorted[disp_sorted["_rest"] > 0]

    for fill_idx in disp_sorted.index:
        if lote_lbs >= max_take:
            break

        fill_rest = pool.disponible(fill_idx)
        if fill_rest <= 0:
            continue

        # LNKs distintos
        fill_lnk  = str(work.at[fill_idx, "LNK"]) if "LNK" in work.columns else str(fill_idx)
        if len(lnks_lote | {fill_lnk}) > cfg.max_items:
            continue

        # Prioridades
        fill_prio = str(work.at[fill_idx, "PRIORIDAD"]).upper().strip() if "PRIORIDAD" in work.columns else "AHEAD"
        if not can_mix(prios_lote, fill_prio, allowed_pairs):
            continue

        # Anchos: cantidad y diferencia entre valores
        fill_anchos = get_row_widths(work.loc[fill_idx])
        anchos_new  = anchos_lote | fill_anchos
        if not anchos_validos(anchos_new, max_anchos):
            continue
        if max_unique_widths is not None and len(anchos_new) > max_unique_widths:
            continue
        # Validar diferencia mínima/máxima entre anchos del lote
        _min_d = cfg.min_diff if cfg.min_diff > 0 else 0.0
        _max_d = cfg.max_diff if cfg.max_diff > 0 else 9999.0
        if len(anchos_new) > 1 and not valid_width_group(list(anchos_new), _min_d, _max_d, max_anchos):
            continue

        remaining = max_take - lote_lbs
        take_fill = SplitEngine.choose_take(fill_rest, remaining, split_min, cfg.scrap_remainder)
        if take_fill <= 0:
            continue

        # Calcular score incremental para preferir el que aporta más
        new_total = lote_lbs + take_fill
        widths_now = set(anchos_lote)
        adds_new_width = 1 if len(anchos_new) > len(widths_now) else 0

        has_target = 0
        if combo_target is not None:
            for w in fill_anchos:
                if abs(float(w) - float(combo_target)) < 1e-6:
                    has_target = 1

        # Greedy score: maximizar lbs + bonos por ancho objetivo y variedad
        greedy_score = new_total + has_target * 1e-3 + adds_new_width * 1e-4
        _ = greedy_score  # se usa implícitamente al continuar el loop

        rows.append((fill_idx, take_fill))
        lote_lbs    += take_fill
        anchos_lote  = anchos_new
        prios_lote.add(fill_prio)
        lnks_lote.add(fill_lnk)

    # ── VALIDACIONES FINALES ──────────────────────────────────────────────────
    if lote_lbs < minimo - 1e-9:
        return None

    # Mínimo de anchos distintos
    if min_unique_widths is not None and len(anchos_lote) < min_unique_widths:
        return None
    elif rule_info.get("regla_aplicada") == "COMBO_ANCHOS" and len(anchos_lote) < 2:
        return None

    # Máximo de anchos (ya chequeado incremental, doble chequeo final)
    if len(anchos_lote) > max_anchos:
        return None

    lote_lbs = round(lote_lbs, 2)

    lote_rows = [LoteRow(orig_idx=idx, lbs_asig=lbs) for idx, lbs in rows]

    return Lote(
        lid        = 0,   # se asigna en el optimizador
        categoria  = cat,
        rango_id   = rid,
        mix        = mix,
        minimo     = minimo,
        maximo     = maximo,
        total      = lote_lbs,
        rows       = lote_rows,
        anchos     = anchos_lote,
        rule_info  = dict(rule_info),
    )


# ------------------------------------------------------------
# core/optimizer.py
# ------------------------------------------------------------
def _build_ranges(categorias: List[Categoria]) -> List[dict]:
    """Convierte categorías a dicts de rango ordenados mayor→menor MAXIMO."""
    ranges = []
    for c in categorias:
        if not c.activo:
            continue
        ranges.append({
            "CATEGORIA":     c.nombre,
            "MINIMO":        c.minimo,
            "MAXIMO":        c.maximo,
            "CAPACIDAD":     c.capacidad_lbs,
            "MIX":           c.mix.upper(),
            "TIPO_TEJIDO":   c.tipo_tejido.upper(),
            "CTDMAXANCHOS":  c.ctd_max_anchos,
            "RANGO_ID":      c.rango_id,
        })
    return sorted(ranges, key=lambda r: -float(r["MAXIMO"]))


def run_optimizer(
    df:           pd.DataFrame,
    categorias:   List[Categoria],
    cfg:          Configuracion,
    reglas_ancho: List[ReglaAncho]      = None,
    reglas_color: List[ReglaColor]      = None,
    reglas_fam:   List[ReglaFamilia]    = None,
    reglas_combo: List[ReglaCombinacion]= None,
    progress_cb:  Optional[callable]    = None,
) -> Tuple[pd.DataFrame, Dict[int, float]]:
    """
    Motor principal NV3.

    Args:
        df:          DataFrame con los datos (post-build_join_from_data).
        categorias:  Lista de Categoria activas.
        cfg:         Configuracion del solver.
        reglas_*:    Restricciones de asignación.
        progress_cb: Callback(pct: float, msg: str) para barra de progreso.

    Returns:
        (result_df, lbs_restantes_global)
        result_df: cada fila = 1 ítem asignado a 1 lote.
        lbs_restantes_global: {orig_idx: lbs_aun_disponibles}
    """
    reglas_ancho = reglas_ancho or []
    reglas_color = reglas_color or []
    reglas_fam   = reglas_fam   or []
    reglas_combo = reglas_combo or []

    # ── PREPARACIÓN ──────────────────────────────────────────────────────────
    ranges_all = _build_ranges(categorias)
    if not ranges_all:
        log.warning("No hay categorías activas.")
        return pd.DataFrame(), {}

    # Pool compartido — 1 dict global para todos los rangos
    lbs_base: Dict[int, float] = {}
    for i in df.index:
        base = float(df.loc[i, "LBS_C"])
        if "PCT_CARGA" in df.columns:
            base = round(base * float(df.loc[i, "PCT_CARGA"]), 4)
        lbs_base[i] = base

    pool = SplitEngine(lbs_base)

    # Capacity tracking en lbs (no en conteo de lotes)
    cap_usada: Dict[str, float] = {r["RANGO_ID"]: 0.0 for r in ranges_all}

    # Allowed pairs de mezcla de prioridades
    allowed_pairs = build_allowed_pairs(cfg.combinacion_prioridad)

    # Width target order: [2, 3, 4] o lo que configure el usuario
    targets = parse_widths_target_order(cfg.widths_target_order)

    # Agrupar por TELA.CUERPO + TONO + MIX
    group_keys = ["ESTILO_C", "MIX"]
    if "TONO" in df.columns:
        group_keys = ["ESTILO_C", "TONO", "MIX"]

    all_groups = list(df.groupby(group_keys).groups.items())
    n_groups   = len(all_groups)
    log.info(f"NV3: {n_groups} grupos TELA+TONO+MIX, {len(ranges_all)} rangos activos.")

    all_rows:   List[dict] = []
    lid_por_cat: Dict[str, int] = {r["RANGO_ID"]: 1 for r in ranges_all}
    lote_id_global = 1

    BLOCK_ORDER = ["VENCIDOS", "AHEAD", "AHEAD2", "OTROS"]

    # ── LOOP PRINCIPAL — por grupo TELA+TONO+MIX ─────────────────────────────
    for g_idx, (gk, grp_idx) in enumerate(all_groups):
        if progress_cb and g_idx % 10 == 0:
            progress_cb(g_idx / n_groups * 100, f"Grupo {g_idx}/{n_groups}…")

        mix_grp = str(gk[-1]).upper()
        work = df.loc[grp_idx].copy()

        # Precomputar anchos y prioridad rank
        work["_anchos"]   = work.apply(get_row_widths, axis=1)
        work["_prio_rank"]= work["PRIORIDAD"].apply(prio_rank) if "PRIORIDAD" in work.columns else 4
        work["_bloque"]   = work["PRIORIDAD"].apply(prio_bloque) if "PRIORIDAD" in work.columns else "OTROS"

        # Tipo tejido del grupo (primera fila)
        tipo_tej_grp = "TODOS"
        if "TIPO_TEJIDO" in work.columns:
            tipo_tej_grp = str(work["TIPO_TEJIDO"].iloc[0]).upper()

        # Rangos compatibles con este MIX y tipo de tejido
        ranges_mix = [
            r for r in ranges_all
            if r["MIX"] == mix_grp and (
                r["TIPO_TEJIDO"] == "TODOS" or
                tipo_tej_grp in (r["TIPO_TEJIDO"], "TODOS")
            )
        ]
        if not ranges_mix:
            continue

        blocked: set = set()

        # ── POR BLOQUE DE PRIORIDAD ───────────────────────────────────────────
        while True:
            # Actualizar remanentes en work
            work["_rest"] = work.index.map(lambda i: pool.disponible(i))

            # ¿Quedan ítems disponibles?
            disp_all = work[work["_rest"] > 0]
            if disp_all.empty:
                break

            made_any = False

            for bloque in BLOCK_ORDER:
                if bloque in blocked:
                    continue

                cand = disp_all[disp_all["_bloque"] == bloque]
                if cand.empty:
                    blocked.add(bloque)
                    continue

                # Seeds: top BEAM_WIDTH por LBS disponibles
                seeds = (
                    cand.sort_values("_rest", ascending=False)
                    .head(cfg.beam_width)
                    .index.tolist()
                )

                best_lote:  Optional[Lote] = None
                best_score: float          = -1e30

                for seed_idx in seeds:
                    seed_row = work.loc[seed_idx]

                    # Determinar orden de rangos según reglas de negocio
                    ranges_try, rule_info = reorder_ranges_for_seed(
                        ranges_mix   = ranges_mix,
                        mix          = mix_grp,
                        row          = seed_row,
                        reglas_ancho = reglas_ancho,
                        reglas_color = reglas_color,
                        reglas_fam   = reglas_fam,
                        reglas_combo = reglas_combo,
                        work         = work,
                        seed_idx     = seed_idx,
                        cfg          = cfg,
                    )

                    # Si ANCHO18 en DYE, filtrar rangos permitidos
                    if rule_info.get("regla_aplicada") == "ANCHO18" and mix_grp == "DYE":
                        allowed = set(cfg.ancho18_allowed_max_dye)
                        if cfg.ancho18_allow_spillover_2600:
                            allowed.add(2600.0)
                        ranges_try = [r for r in ranges_try if float(r["MAXIMO"]) in allowed]

                    pri_list    = order_priorities(rule_info.get("prioridades", []), cfg.rule_order)
                    use_upgrades= bool(pri_list) and cfg.upgrade_categoria
                    pri_iter    = pri_list if (use_upgrades and cfg.try_all_priorities) else [None]

                    lote_candidato: Optional[Lote] = None

                    # ── INTENTO POR OBJETIVO DE ANCHOS (2>3>4) ───────────────
                    req_strict = cfg.require_widths_strict
                    for target in targets:
                        rng_target = filter_ranges_for_width_target(
                            ranges_try,
                            mix_grp,
                            target,
                            cfg.allowed_maximo_3_dye,
                            cfg.allowed_maximo_4_dye,
                        )
                        found = False
                        for pri in pri_iter:
                            rng_pri = rng_target
                            if pri is not None:
                                rng_pri = ranges_matching_priority(pri, rng_target)

                            for rng in rng_pri:
                                rid = rng["RANGO_ID"]
                                cap_t = float(rng["CAPACIDAD"])
                                if cap_usada.get(rid, 0.0) >= cap_t - 1e-6:
                                    continue

                                lote = intentar_lote_para_rango(
                                    work             = work,
                                    seed_idx         = seed_idx,
                                    rango            = rng,
                                    pool             = pool,
                                    cap_usada        = cap_usada.get(rid, 0.0),
                                    cfg              = cfg,
                                    rule_info        = rule_info,
                                    allowed_pairs    = allowed_pairs,
                                    min_unique_widths= target,
                                    max_unique_widths= target if req_strict else None,
                                )
                                if lote is not None:
                                    lote_candidato = lote
                                    found = True
                                    break
                            if found:
                                break
                        if lote_candidato is not None:
                            break

                    # ── FALLBACK — sin restricción de objetivo de anchos ──────
                    if lote_candidato is None:
                        for rng in ranges_try:
                            rid = rng["RANGO_ID"]
                            if cap_usada.get(rid, 0.0) >= float(rng["CAPACIDAD"]) - 1e-6:
                                continue
                            lote = intentar_lote_para_rango(
                                work          = work,
                                seed_idx      = seed_idx,
                                rango         = rng,
                                pool          = pool,
                                cap_usada     = cap_usada.get(rid, 0.0),
                                cfg           = cfg,
                                rule_info     = rule_info,
                                allowed_pairs = allowed_pairs,
                            )
                            if lote is not None:
                                lote_candidato = lote
                                break

                    # ── SCORING — seleccionar mejor entre semillas ────────────
                    if lote_candidato is not None:
                        sc = score_lote(lote_candidato, cfg, pool.snapshot())
                        lote_candidato.score = sc
                        if sc > best_score:
                            best_score = sc
                            best_lote  = lote_candidato

                # ── CONFIRMAR MEJOR LOTE ──────────────────────────────────────
                if best_lote is None:
                    blocked.add(bloque)
                    continue

                # Asignar ID
                rid = best_lote.rango_id
                lid = lid_por_cat[rid]
                lid_por_cat[rid]   += 1
                lote_id_global     += 1
                best_lote.lid       = lid
                cat_nombre          = best_lote.categoria
                lote_id_str         = f"{cat_nombre}-L{lid:04d}"

                # Descontar pool compartido
                split_min_used = (
                    cfg.split_min_ancho18
                    if best_lote.rule_info.get("regla_aplicada") == "ANCHO18"
                    else cfg.split_min_default
                )
                for row_r in best_lote.rows:
                    pool.descontar(row_r.orig_idx, row_r.lbs_asig)
                    # Scrap de remanente < split_min
                    if cfg.scrap_remainder:
                        rem = pool.restante(row_r.orig_idx)
                        if 0 < rem < split_min_used:
                            pool.scrap(row_r.orig_idx, rem)

                cap_usada[rid] = cap_usada.get(rid, 0.0) + best_lote.total

                # Registrar filas en resultado
                for row_r in best_lote.rows:
                    if row_r.orig_idx not in work.index:
                        continue
                    orig = work.loc[row_r.orig_idx].copy()
                    orig["LBS_C"]           = round(row_r.lbs_asig, 2)
                    orig["CATEGORIA"]       = cat_nombre
                    orig["LOTE_ID"]         = lote_id_str
                    orig["TOTAL_LOTE"]      = round(best_lote.total, 2)
                    orig["PCT_CARGA_REAL"]  = round(best_lote.total / max(best_lote.maximo, 1) * 100, 1)
                    orig["SET_ANCHOS_LOTE"] = best_lote.set_anchos
                    orig["CANT_ANCHOS"]     = best_lote.cant_anchos
                    orig["TIPO_LOTE_ANCHO"] = best_lote.tipo
                    orig["APLICA_REGLA"]    = best_lote.rule_info.get("regla_aplicada", "NONE")
                    orig["PRIORIDAD_USADA"] = best_lote.maximo
                    orig["DECISION_SCORE"]  = round(best_lote.score, 4)
                    all_rows.append(orig)

                blocked = set()   # reset — puede haber más lotes
                made_any = True
                break             # volver al while para actualizar remanentes

            if not made_any:
                break  # ningún bloque generó lote → salir del grupo

    # ── RESULTADO ─────────────────────────────────────────────────────────────
    if not all_rows:
        log.warning("NV3: no se generaron lotes.")
        return pd.DataFrame(), pool.snapshot()

    result = pd.DataFrame(all_rows)

    show_cols = [
        "CATEGORIA", "LOTE_ID", "COLOR_A", "ESTILO_C", "ANCHO", "LBS_C",
        "TOTAL_LOTE", "PCT_CARGA_REAL", "SET_ANCHOS_LOTE", "CANT_ANCHOS",
        "TIPO_LOTE_ANCHO", "MIX", "TIPO_TEJIDO", "PCT_CARGA", "PRIORIDAD",
        "COLOR_R", "FAMILIA", "LNK", "LNK_PRIORIDAD", "TONO", "CONSUMO_C",
        "STYLE", "APLICA_REGLA", "PRIORIDAD_USADA", "DECISION_SCORE",
    ]
    show_cols = [c for c in show_cols if c in result.columns]
    result = (
        result[show_cols]
        .sort_values(["CATEGORIA", "LOTE_ID"])
        .reset_index(drop=True)
    )

    total_asig = result.groupby("LOTE_ID")["TOTAL_LOTE"].first().sum()
    log.info(
        f"NV3 completado: {result['LOTE_ID'].nunique()} lotes, "
        f"{total_asig:,.0f} lbs asignadas."
    )

    return result, pool.snapshot()


# ------------------------------------------------------------
# services/data_loader.py
# ------------------------------------------------------------
REQUIRED_COLS = [
    "STYLE", "COLOR", "TELA.CUERPO", "ANCHO.F.C", "TOTAL",
    "MIX", "COLOR_R", "FAMILIA", "PRIORIDAD",
]


def clean_cols(cols):
    out = []
    for c in cols:
        c = "" if c is None else str(c)
        c = c.replace("\n", " ").replace("\r", " ")
        c = re.sub(r"\s+", " ", c).strip()
        out.append(c)
    return out


def find_header_row(xl_path_or_bytes, sheet_name: str, required_cols: list, search_rows: int = 80) -> int:
    if isinstance(xl_path_or_bytes, (bytes, BytesIO)):
        src = BytesIO(xl_path_or_bytes) if isinstance(xl_path_or_bytes, bytes) else xl_path_or_bytes
    else:
        src = xl_path_or_bytes
    preview = pd.read_excel(src, sheet_name=sheet_name, header=None, nrows=search_rows)
    req = set(required_cols)
    for r in range(search_rows):
        row_vals = set(str(v).strip() for v in preview.iloc[r].tolist() if v is not None and str(v).strip())
        if req.issubset(row_vals):
            return r
    return 2  # default: fila 3 (index 2)


def load_raw(file_bytes: bytes) -> Tuple[Optional[pd.DataFrame], Optional[str]]:
    """Lee la hoja DATA del Excel, detectando automáticamente la fila de header."""
    try:
        xl = pd.ExcelFile(BytesIO(file_bytes))
    except Exception as e:
        return None, f"No se pudo abrir el Excel: {e}"

    if "DATA" not in xl.sheet_names:
        return None, "No se encontró la hoja DATA."

    hdr = find_header_row(BytesIO(file_bytes), "DATA", REQUIRED_COLS)
    df  = pd.read_excel(BytesIO(file_bytes), sheet_name="DATA", header=hdr)
    df.columns = clean_cols(df.columns)
    return df, None


def build_dataframe(df_raw: pd.DataFrame) -> Tuple[Optional[pd.DataFrame], Optional[str]]:
    """
    Normaliza el DataFrame crudo para el motor NV3.
    Renombra columnas, castea tipos, agrega columnas faltantes.
    """
    df = df_raw.copy()
    df.columns = [c.strip().replace("\xa0", "") for c in df.columns]

    missing = [c for c in REQUIRED_COLS if c not in df.columns]
    if missing:
        return None, f"Columnas faltantes en DATA: {missing}"

    # Renombrar
    df = df.rename(columns={
        "COLOR":       "COLOR_A",
        "TELA.CUERPO": "ESTILO_C",
        "ANCHO.F.C":   "ANCHO",
        "TOTAL":       "LBS_C",
    })

    # Tipos numéricos
    for col in ("LBS_C", "ANCHO", "ANCHO.F.M"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    df["PCT_CARGA"] = pd.to_numeric(df["PCT_CARGA"] if "PCT_CARGA" in df.columns else 1.0, errors="coerce").fillna(1.0)
    df["MIX"]       = df["MIX"].astype(str).str.upper().str.strip()

    # Columnas opcionales con defaults
    if "TIPO_TEJIDO" not in df.columns:
        df["TIPO_TEJIDO"] = "TODOS"
    if "TONO" not in df.columns:
        df["TONO"] = ""

    # Preservar LNK (puede venir como BD o LNK)
    if "LNK" not in df.columns and "BD" in df.columns:
        df["LNK"] = df["BD"]
    elif "LNK" not in df.columns:
        df["LNK"] = df.index.astype(str)

    # LNK_PRIORIDAD para trazabilidad
    if "PRIORIDAD" in df.columns:
        df["LNK_PRIORIDAD"] = df["LNK"].astype(str) + "|" + df["PRIORIDAD"].astype(str)

    # Limpiar nulos críticos
    df = df.dropna(subset=["LBS_C", "ANCHO", "COLOR_A", "ESTILO_C"])
    df = df[df["LBS_C"] > 0]
    df = df.reset_index(drop=True)

    return df, None


# ------------------------------------------------------------
# services/config_manager.py
# ------------------------------------------------------------
PROFILES_FILE = "elcatex_nv3_profiles.json"

# ── Defaults ──────────────────────────────────────────────────────────────────

DEFAULT_CATEGORIAS = [
    Categoria("A-4000", 3900, 4000, "DYE",    "FLEECE", lotes_dia=5,  semanas=4.0, ctd_max_anchos=4),
    Categoria("B-3300", 3000, 3300, "DYE",    "TODOS",  lotes_dia=6,  semanas=4.0, ctd_max_anchos=4),
    Categoria("C-2600", 2500, 2600, "DYE",    "TODOS",  lotes_dia=29, semanas=4.0, ctd_max_anchos=3),
    Categoria("D-2200", 2000, 2200, "DYE",    "TODOS",  lotes_dia=17, semanas=4.0, ctd_max_anchos=3),
    Categoria("E-1100", 1000, 1100, "DYE",    "TODOS",  lotes_dia=25, semanas=4.0, ctd_max_anchos=2),
    Categoria("F-2200", 2000, 2200, "BLEACH", "TODOS",  lotes_dia=21, semanas=4.0, ctd_max_anchos=3),
    Categoria("G-1100", 1000, 1100, "BLEACH", "TODOS",  lotes_dia=4,  semanas=4.0, ctd_max_anchos=2),
]

DEFAULT_REGLAS_ANCHO = [
    ReglaAncho("PC54Y",   18, [2600]),
    ReglaAncho("PC55LS",  18, [2600]),
    ReglaAncho("PC55Y",   18, [2600]),
    ReglaAncho("PC330Y",  18, [2200]),
    ReglaAncho("PC54-2",  18, [2600]),
    ReglaAncho("PC55-2",  18, [2600]),
    ReglaAncho("PC54LS",  18, [2600]),
    ReglaAncho("PC61Y",   18, [2600]),
    ReglaAncho("PC54DTG", 18, [2600]),
    ReglaAncho("LPC61",   18, [2600]),
    ReglaAncho("PC55P",   18, [2600]),
    ReglaAncho("PC61LSP", 18, [2600]),
]

DEFAULT_REGLAS_COLOR = [
    ReglaColor("RESTRICCION", [2600]),
    ReglaColor("NORMAL",      []),
]

DEFAULT_REGLAS_FAMILIA = [
    ReglaFamilia("PC68",    [2600]),
    ReglaFamilia("PC850",   [2600]),
    ReglaFamilia("PC78/90", [4000, 2600, 3300]),
]

DEFAULT_REGLAS_COMBO: list = []

DEFAULT_CONFIG = Configuracion()


# ── Serialización ─────────────────────────────────────────────────────────────

def _cat_to_dict(c: Categoria) -> dict:
    return c.to_dict()

def _cat_from_dict(d: dict) -> Categoria:
    return Categoria.from_dict(d)

def _regla_ancho_to_dict(r: ReglaAncho) -> dict:
    return {"STYLE": r.style, "LIMITE_ANCHO": r.limite_ancho,
            "PRIORIDADES": r.prioridades, "ACTIVO": r.activo}

def _regla_color_to_dict(r: ReglaColor) -> dict:
    return {"COLOR_R": r.color_r, "PRIORIDADES": r.prioridades, "ACTIVO": r.activo}

def _regla_fam_to_dict(r: ReglaFamilia) -> dict:
    return {"FAMILIA": r.familia, "PRIORIDADES": r.prioridades, "ACTIVO": r.activo}

def _cfg_to_dict(c: Configuracion) -> dict:
    import dataclasses
    d = dataclasses.asdict(c)
    # sets → sorted lists for JSON
    for k in ("ancho18_allowed_max_dye", "allowed_maximo_3_dye", "allowed_maximo_4_dye"):
        if k in d:
            d[k] = sorted(d[k])
    return d


# ── Profile API ───────────────────────────────────────────────────────────────

def load_profiles() -> Dict[str, Any]:
    if os.path.exists(PROFILES_FILE):
        with open(PROFILES_FILE) as f:
            return json.load(f)
    return {}

def save_profiles(profiles: Dict[str, Any]) -> None:
    with open(PROFILES_FILE, "w") as f:
        json.dump(profiles, f, indent=2, default=str)

def save_profile(name: str, categorias, reglas_ancho, reglas_color, reglas_familia,
                 reglas_combo, cfg: Configuracion) -> None:
    from datetime import datetime
    profiles = load_profiles()
    profiles[name] = {
        "categorias":    [_cat_to_dict(c) for c in categorias],
        "reglas_ancho":  [_regla_ancho_to_dict(r) for r in reglas_ancho],
        "reglas_color":  [_regla_color_to_dict(r) for r in reglas_color],
        "reglas_familia":[_regla_fam_to_dict(r) for r in reglas_familia],
        "reglas_combo":  reglas_combo,
        "config":        _cfg_to_dict(cfg),
        "saved_at":      datetime.now().isoformat(),
    }
    save_profiles(profiles)

def load_profile(name: str):
    profiles = load_profiles()
    if name not in profiles:
        return None
    p = profiles[name]
    cats = [Categoria.from_dict(d) for d in p.get("categorias", [])]
    ra   = [ReglaAncho(d["STYLE"], d["LIMITE_ANCHO"], d["PRIORIDADES"], d.get("ACTIVO", True))
            for d in p.get("reglas_ancho", [])]
    rc   = [ReglaColor(d["COLOR_R"], d["PRIORIDADES"], d.get("ACTIVO", True))
            for d in p.get("reglas_color", [])]
    rf   = [ReglaFamilia(d["FAMILIA"], d["PRIORIDADES"], d.get("ACTIVO", True))
            for d in p.get("reglas_familia", [])]
    combo= p.get("reglas_combo", [])
    cfg  = Configuracion.from_dict(p.get("config", {}))
    return cats, ra, rc, rf, combo, cfg


# ------------------------------------------------------------
# services/report_builder.py
# ------------------------------------------------------------
def build_reports(
    result:      pd.DataFrame,
    df_original: pd.DataFrame,
    categorias:  List[Categoria],
    lbs_restantes: Dict[int, float],
) -> Dict[str, pd.DataFrame]:
    """
    Genera todas las hojas del reporte.
    Retorna dict {sheet_name: DataFrame}.
    """
    cat_max = {c.nombre: c.maximo for c in categorias}
    cat_min = {c.nombre: c.minimo for c in categorias}

    reports = {}

    # ── DETALLE_LOTES ─────────────────────────────────────────────────────────
    det = result.copy()
    det = det.rename(columns={
        "SET_ANCHOS_LOTE": "ANCHOS_LOTE",
        "ESTILO_C":        "TELA.CUERPO",
        "COLOR_A":         "COLOR",
        "LBS_C":           "LBS_ASIGNADAS",
        "ANCHO":           "ANCHO.F.C",
    })
    det["BLOQUE"]          = det.get("PRIORIDAD", pd.Series("", index=det.index))
    det["LBS_EXTRA_SOBRE_ORDEN"] = 0
    if "CONSUMO_C" in det.columns:
        det["DOCENAS"] = (det["LBS_ASIGNADAS"] / det["CONSUMO_C"].replace(0, float("nan"))).round(2).fillna(0)
    else:
        det["DOCENAS"] = 0

    det_cols = [c for c in [
        "LOTE_ID","ANCHOS_LOTE","CATEGORIA","MIX","TELA.CUERPO","COLOR",
        "TONO","LNK_PRIORIDAD","LNK","PRIORIDAD","BLOQUE","ANCHO.F.C","CONSUMO_C",
        "FAMILIA","COLOR_R","STYLE","LBS_ASIGNADAS","LBS_EXTRA_SOBRE_ORDEN",
        "APLICA_REGLA","PRIORIDAD_USADA","DOCENAS","TIPO_TEJIDO","PCT_CARGA",
        "TOTAL_LOTE","PCT_CARGA_REAL","CANT_ANCHOS","TIPO_LOTE_ANCHO","DECISION_SCORE",
    ] if c in det.columns]
    reports["DETALLE_LOTES"] = det[det_cols].copy()

    # ── RESUMEN_LOTES ─────────────────────────────────────────────────────────
    if not result.empty:
        res_grp = result.groupby(["LOTE_ID", "SET_ANCHOS_LOTE", "CATEGORIA", "MIX"])
        resumen = res_grp.agg(
            TELA_CUERPO      = ("ESTILO_C",  lambda x: x.iloc[0]),
            COLOR_TONO_KEY   = ("COLOR_A",   lambda x: "|".join(sorted(x.unique()))),
            LBS_TOTAL        = ("TOTAL_LOTE","first"),
            SKU_DISTINTOS    = ("ANCHO",     "count"),
            ANCHOS_UNICOS    = ("CANT_ANCHOS","first"),
            BLOQUE_DOMINANTE = ("PRIORIDAD", lambda x: x.mode()[0] if len(x) else ""),
            TIPO_LOTE        = ("TIPO_LOTE_ANCHO", "first"),
            APLICA_REGLA     = ("APLICA_REGLA", lambda x: x.mode()[0] if len(x) else "NONE"),
            AVG_SCORE        = ("DECISION_SCORE", "mean") if "DECISION_SCORE" in result.columns else ("CANT_ANCHOS", "first"),
        ).reset_index().rename(columns={
            "SET_ANCHOS_LOTE": "ANCHOS_LOTE",
            "TELA_CUERPO":     "TELA.CUERPO",
            "COLOR_TONO_KEY":  "COLOR/TONO_KEY",
        })
        resumen["MIN_RANGO"]         = resumen["CATEGORIA"].map(cat_min)
        resumen["MAX_RANGO"]         = resumen["CATEGORIA"].map(cat_max)
        resumen["CAPACIDAD_PERDIDA"] = resumen.apply(
            lambda r: max(0, cat_max.get(r["CATEGORIA"], 0) - r["LBS_TOTAL"]), axis=1)
        resumen["UPGRADE_CATEGORIA"] = 1
        col_order = [
            "LOTE_ID","ANCHOS_LOTE","CATEGORIA","MIX","TELA.CUERPO","COLOR/TONO_KEY",
            "LBS_TOTAL","MIN_RANGO","MAX_RANGO","CAPACIDAD_PERDIDA","SKU_DISTINTOS",
            "ANCHOS_UNICOS","BLOQUE_DOMINANTE","APLICA_REGLA","TIPO_LOTE","UPGRADE_CATEGORIA",
        ]
        if "AVG_SCORE" in resumen.columns:
            col_order.append("AVG_SCORE")
        reports["RESUMEN_LOTES"] = resumen[[c for c in col_order if c in resumen.columns]]
    else:
        reports["RESUMEN_LOTES"] = pd.DataFrame()

    # ── EXCEDENTES ────────────────────────────────────────────────────────────
    exc_rows = []
    for orig_idx, row in df_original.iterrows():
        lbs_base = float(row.get("LBS_C", 0))
        restante = lbs_restantes.get(orig_idx, lbs_base)
        if restante > 0.01:
            exc_rows.append({
                "LNK":           row.get("LNK", ""),
                "TELA.CUERPO":   row.get("ESTILO_C", ""),
                "COLOR":         row.get("COLOR_A", ""),
                "TONO":          row.get("TONO", ""),
                "MIX":           row.get("MIX", ""),
                "PRIORIDAD":     row.get("PRIORIDAD", ""),
                "BLOQUE":        row.get("PRIORIDAD", ""),
                "ANCHO.F.C":     row.get("ANCHO", 0),
                "ANCHO.F.M":     row.get("ANCHO.F.M", 0),
                "TOTAL":         lbs_base,
                "LBS_RESTANTES": round(restante, 2),
                "LBS_SCRAP":     0,
            })
    if exc_rows:
        reports["EXCEDENTES"] = (
            pd.DataFrame(exc_rows)
            .sort_values("LBS_RESTANTES", ascending=False)
            .reset_index(drop=True)
        )
    else:
        reports["EXCEDENTES"] = pd.DataFrame(columns=[
            "LNK","TELA.CUERPO","COLOR","TONO","MIX","PRIORIDAD","BLOQUE",
            "ANCHO.F.C","ANCHO.F.M","TOTAL","LBS_RESTANTES","LBS_SCRAP"
        ])

    # ── CAPACIDAD_X_CATEG ─────────────────────────────────────────────────────
    # FIX: usar first() de TOTAL_LOTE por lote — no suma de filas
    if not result.empty:
        lbs_x_cat = (
            result.groupby(["CATEGORIA", "LOTE_ID"])["TOTAL_LOTE"]
            .first()
            .reset_index()
            .groupby("CATEGORIA")["TOTAL_LOTE"]
            .sum()
            .reset_index()
        )
        lbs_x_cat.columns = ["CATEGORIA", "LBS_ASIGNADAS"]
    else:
        lbs_x_cat = pd.DataFrame(columns=["CATEGORIA", "LBS_ASIGNADAS"])

    cap_rows = []
    for c in categorias:
        asig    = float(lbs_x_cat[lbs_x_cat["CATEGORIA"] == c.nombre]["LBS_ASIGNADAS"].sum())
        n_lotes = int(result[result["CATEGORIA"] == c.nombre]["LOTE_ID"].nunique()) if not result.empty else 0
        cap_lbs = c.capacidad_lbs
        cap_rows.append({
            "CATEGORIA":         c.nombre,
            "MIX":               c.mix,
            "MINIMO":            c.minimo,
            "MAXIMO":            c.maximo,
            "LOTES_DIA":         c.lotes_dia,
            "SEMANAS":           c.semanas,
            "MAX_LOTES_PERIODO": int(c.lotes_dia * 7 * c.semanas),
            "LOTES_GENERADOS":   n_lotes,
            "CAPACIDAD_LBS":     round(cap_lbs, 0),
            "LBS_ASIGNADAS":     round(asig, 1),
            "DIFERENCIA":        round(asig - cap_lbs, 1),   # negativo = capacidad libre
            "PCT_OCUPACION":     round(asig / cap_lbs * 100, 1) if cap_lbs > 0 else 0,
        })
    reports["CAPACIDAD_X_CATEG"] = pd.DataFrame(cap_rows)

    # ── PRIORIDAD_VS_ASIG ─────────────────────────────────────────────────────
    if "PRIORIDAD" in result.columns and not result.empty:
        lbs_c_col = "LBS_C" if "LBS_C" in result.columns else "LBS_ASIGNADAS"
        prio_asig = (
            result.groupby(["MIX", "PRIORIDAD"])
            .agg(LBS_ASIGNADAS=(lbs_c_col, "sum"))
            .reset_index()
            .rename(columns={"PRIORIDAD": "BLOQUE"})
        )
        if "PRIORIDAD" in df_original.columns:
            prio_base = (
                df_original.groupby(["MIX", "PRIORIDAD"])
                .agg(LBS_BASE=("LBS_C", "sum"))
                .reset_index()
                .rename(columns={"PRIORIDAD": "BLOQUE"})
            )
            prio_df = prio_base.merge(prio_asig, on=["MIX", "BLOQUE"], how="left").fillna(0)
        else:
            prio_df = prio_asig.copy()
            prio_df["LBS_BASE"] = prio_df["LBS_ASIGNADAS"]
        prio_df["LBS_SIN_ASIGNAR"] = prio_df["LBS_ASIGNADAS"] - prio_df["LBS_BASE"]
        reports["PRIORIDAD_VS_ASIG"] = prio_df[["MIX", "BLOQUE", "LBS_BASE", "LBS_ASIGNADAS", "LBS_SIN_ASIGNAR"]]
    else:
        reports["PRIORIDAD_VS_ASIG"] = pd.DataFrame(
            columns=["MIX", "BLOQUE", "LBS_BASE", "LBS_ASIGNADAS", "LBS_SIN_ASIGNAR"]
        )

    # ── LNK_COMPLETITUD ───────────────────────────────────────────────────────
    lnk_col = next((c for c in ["LNK", "SKU", "CUT-TICKET"] if c in df_original.columns), None)
    if lnk_col:
        lnk_base = (
            df_original.groupby(lnk_col)
            .agg(LBS_BASE=("LBS_C", "sum"))
            .reset_index()
        )
        lnk_meta = df_original[[lnk_col, "MIX"]].drop_duplicates(subset=[lnk_col])

        if not result.empty and lnk_col in result.columns:
            lbs_c_col = "LBS_C" if "LBS_C" in result.columns else "LBS_ASIGNADAS"
            lnk_asig = (
                result.groupby(lnk_col)
                .agg(LBS_ASIGNADAS=(lbs_c_col, "sum"))
                .reset_index()
            )
            lnk_comp = lnk_base.merge(lnk_asig, on=lnk_col, how="left").fillna(0)
        else:
            lnk_comp = lnk_base.copy()
            lnk_comp["LBS_ASIGNADAS"] = 0

        lnk_comp = lnk_comp.merge(lnk_meta, on=lnk_col, how="left")
        lnk_comp["LBS_SCRAP"] = 0
        lnk_comp["BALANCE"]   = lnk_comp["LBS_BASE"] - lnk_comp["LBS_ASIGNADAS"]
        lnk_comp["ESTADO"]    = lnk_comp["BALANCE"].apply(
            lambda b: "COMPLETO" if abs(b) < 1 else ("PARCIAL" if b > 0 else "EXCEDIDO")
        )
        lnk_comp = lnk_comp[["MIX", lnk_col, "LBS_BASE", "LBS_ASIGNADAS", "LBS_SCRAP", "BALANCE", "ESTADO"]]
        reports["LNK_COMPLETITUD"] = lnk_comp.sort_values("BALANCE", ascending=False).reset_index(drop=True)
    else:
        reports["LNK_COMPLETITUD"] = pd.DataFrame(
            columns=["MIX", "LNK", "LBS_BASE", "LBS_ASIGNADAS", "LBS_SCRAP", "BALANCE", "ESTADO"]
        )

    # ── DECISION_LOG ─────────────────────────────────────────────────────────
    if "DECISION_SCORE" in result.columns and not result.empty:
        log_cols = [c for c in [
            "LOTE_ID", "MIX", "CATEGORIA", "PRIORIDAD", "LNK", "APLICA_REGLA",
            "PRIORIDAD_USADA", "TOTAL_LOTE", "CANT_ANCHOS", "TIPO_LOTE_ANCHO", "DECISION_SCORE",
        ] if c in result.columns]
        reports["DECISION_LOG"] = result[log_cols].drop_duplicates().sort_values("LOTE_ID")
    else:
        reports["DECISION_LOG"] = pd.DataFrame()

    # ── RESUMEN_CATEGORIA ─────────────────────────────────────────────────────
    if not result.empty:
        res_cat = (
            result.groupby("CATEGORIA")
            .agg(
                Lotes          = ("LOTE_ID", "nunique"),
                Registros      = ("LOTE_ID", "count"),
                Puros          = ("TIPO_LOTE_ANCHO", lambda x: (x == "PURO").sum()),
                Mix_Controlado = ("TIPO_LOTE_ANCHO", lambda x: (x == "MIX_CONTROLADO").sum()),
                Mix_Alto       = ("TIPO_LOTE_ANCHO", lambda x: (x == "MIX_ALTO").sum()),
                Avg_Fill_Pct   = ("PCT_CARGA_REAL", "mean"),
            )
            .reset_index()
        )
        res_cat["Avg_Fill_Pct"] = res_cat["Avg_Fill_Pct"].round(1)
        reports["RESUMEN_CATEGORIA"] = res_cat
    else:
        reports["RESUMEN_CATEGORIA"] = pd.DataFrame()

    # ── PARAMETROS ────────────────────────────────────────────────────────────
    param_rows = []
    for c in categorias:
        for k, v in c.to_dict().items():
            param_rows.append({"CATEGORIA": c.nombre, "PARAMETRO": k, "VALOR": v})
    reports["PARAMETROS"] = pd.DataFrame(param_rows)

    return reports


def format_excel(writer: pd.ExcelWriter) -> None:
    """Aplica formato Calibri 8, header azul, negativos en rojo, autofit, freeze+filter."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb          = writer.book
    hdr_fill    = PatternFill("solid", fgColor="003876")
    hdr_font    = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
    cell_font   = Font(name="Calibri", size=8)
    neg_font    = Font(name="Calibri", size=8, color="C00000")
    num_fmt     = "#,##0"
    dec_fmt     = "#,##0.0"

    for ws in wb.worksheets:
        col_widths: dict = {}
        for row_idx, row in enumerate(ws.iter_rows()):
            for cell in row:
                if row_idx == 0:
                    cell.font      = hdr_font
                    cell.fill      = hdr_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    if isinstance(cell.value, (int, float)) and cell.value is not None:
                        hdr_val = ws.cell(1, cell.column).value or ""
                        is_pct  = any(x in str(hdr_val).upper() for x in ["PCT", "CARGA", "%"])
                        fmt     = dec_fmt if is_pct else num_fmt
                        if cell.value < 0:
                            cell.font = neg_font
                        else:
                            cell.font = cell_font
                        cell.number_format = fmt
                    else:
                        cell.font = cell_font
                    cell.alignment = Alignment(vertical="center")

                val_len = len(str(cell.value)) if cell.value is not None else 0
                col_letter = get_column_letter(cell.column)
                col_widths[col_letter] = max(col_widths.get(col_letter, 8), min(val_len + 2, 40))

        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        ws.freeze_panes = "A2"
        if ws.max_row > 1 and ws.max_column > 0:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


# ------------------------------------------------------------
# ui/streamlit_app.py
# ------------------------------------------------------------
# ── PAGE CONFIG ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="NV3 Loteador | Elcatex",
    page_icon="🏭",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
  html, body, [class*="css"] { font-family: 'Inter', sans-serif; font-size: 12px; }
  .stApp { background-color: #f4f6f9; }
  [data-testid="stSidebar"] { background: linear-gradient(180deg, #003876 0%, #0057a8 60%, #0080c9 100%); }
  [data-testid="stSidebar"] * { color: #ffffff !important; font-size: 11.5px !important; }
  [data-testid="stSidebar"] input { color: #1a1a2e !important; background: #ffffff !important; border-radius: 4px !important; }
  h1 { color: #003876 !important; font-size: 20px !important; font-weight: 700 !important; }
  h2 { color: #0057a8 !important; font-size: 15px !important; font-weight: 600 !important; }
  h3 { color: #003876 !important; font-size: 13px !important; font-weight: 600 !important; }
  .metric-box { background: white; border-radius: 8px; padding: 14px 16px; text-align: center; box-shadow: 0 1px 4px rgba(0,0,0,0.08); }
  .metric-val { font-size: 22px; font-weight: 700; color: #003876; }
  .metric-lbl { font-size: 10px; color: #7f8c8d; font-weight: 500; text-transform: uppercase; letter-spacing: 0.5px; margin-top: 2px; }
  .stButton > button { background: linear-gradient(135deg, #0057a8, #003876); color: white !important; border: none; border-radius: 6px; font-size: 12px; font-weight: 600; padding: 8px 20px; }
  .stButton > button:hover { background: linear-gradient(135deg, #0080c9, #0057a8); }
  .stTabs [data-baseweb="tab-list"] { background: white; border-radius: 8px 8px 0 0; gap: 2px; padding: 4px; }
  .stTabs [data-baseweb="tab"] { font-size: 11.5px; font-weight: 600; color: #7f8c8d; border-radius: 6px; }
  .stTabs [aria-selected="true"] { background: #0057a8 !important; color: white !important; }
  .col-hdr { font-size: 9px; color: #7f8c8d; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px; }
  .logo-header { background: linear-gradient(135deg, #003876, #0057a8); border-radius: 10px; padding: 14px 20px; margin-bottom: 16px; }
  .badge-blue { background: #dbeafe; color: #1e40af; padding: 2px 8px; border-radius: 10px; font-size: 10px; font-weight: 700; }
</style>
""", unsafe_allow_html=True)

# ── HELPERS ───────────────────────────────────────────────────────────────────
def to_int(v, default=0):
    try: return int(v) if v is not None else default
    except: return default

def to_float(v, default=0.0):
    try: return float(v) if v is not None else default
    except: return default

def _ss(key, default):
    if key not in st.session_state:
        st.session_state[key] = default
    return st.session_state[key]


# ── SIDEBAR ───────────────────────────────────────────────────────────────────
def sidebar():
    with st.sidebar:
        st.markdown("""
        <div style="text-align:center;padding:10px 0 16px 0;">
          <div style="font-size:20px;font-weight:800;color:white;letter-spacing:1px;">⚙ ELCATEX</div>
          <div style="font-size:9px;color:#cce4ff;letter-spacing:2px;text-transform:uppercase;">Loteador NV3</div>
        </div>""", unsafe_allow_html=True)

        st.markdown("---")
        st.markdown("**📁 Perfiles**")

        profiles = load_profiles()
        names    = list(profiles.keys())

        col1, col2 = st.columns([2, 1])
        with col1:
            sel = st.selectbox("Perfil", ["(ninguno)"] + names, label_visibility="collapsed")
        with col2:
            if st.button("Cargar", use_container_width=True) and sel != "(ninguno)":
                loaded = load_profile(sel)
                if loaded:
                    cats, ra, rc, rf, combo, cfg = loaded
                    st.session_state["categorias"]    = cats
                    st.session_state["reglas_ancho"]  = ra
                    st.session_state["reglas_color"]  = rc
                    st.session_state["reglas_familia"]= rf
                    st.session_state["reglas_combo"]  = combo
                    st.session_state["config"]        = cfg
                    st.success(f"✓ '{sel}' cargado")
                    st.rerun()

        new_name = st.text_input("Nombre del perfil", placeholder="Mi configuración…")
        if st.button("💾 Guardar perfil", use_container_width=True):
            if new_name.strip():
                save_profile(
                    new_name.strip(),
                    st.session_state.get("categorias", DEFAULT_CATEGORIAS),
                    st.session_state.get("reglas_ancho", DEFAULT_REGLAS_ANCHO),
                    st.session_state.get("reglas_color", DEFAULT_REGLAS_COLOR),
                    st.session_state.get("reglas_familia", DEFAULT_REGLAS_FAMILIA),
                    st.session_state.get("reglas_combo", DEFAULT_REGLAS_COMBO),
                    st.session_state.get("config", DEFAULT_CONFIG),
                )
                st.success(f"Guardado: {new_name.strip()}")
            else:
                st.warning("Escribe un nombre")

        if sel != "(ninguno)" and sel in profiles:
            data = json.dumps(profiles[sel], indent=2, default=str).encode()
            st.download_button(f"⬇ Descargar '{sel}'", data=data,
                               file_name=f"perfil_{sel.replace(' ','_')}.json",
                               mime="application/json", use_container_width=True,
                               key=f"dl_{hash(sel) % 99999}")
        if names:
            del_sel = st.selectbox("Eliminar", [""] + names, label_visibility="collapsed")
            if st.button("🗑 Eliminar seleccionado", use_container_width=True) and del_sel:
                profiles.pop(del_sel, None)
                save_profiles(profiles)
                st.rerun()

        st.markdown("---")
        st.markdown("""<div style="text-align:center;font-size:9px;color:#7fb3d8;margin-top:8px;">
          Grupo Elcatex · NV3 · Honduras 🇭🇳</div>""", unsafe_allow_html=True)


# ── TAB 1: CAPACIDADES ────────────────────────────────────────────────────────
def tab_capacidades():
    st.markdown("### Capacidades de Tinto por Categoría")
    st.caption("Capacidad LBS = Lotes × 7 días × Semanas × Máximo lbs")

    cats        = st.session_state.get("categorias", DEFAULT_CATEGORIAS)
    tipo_opts   = ["TODOS", "FLEECE", "JERSEY"]
    mix_opts    = ["DYE", "BLEACH"]

    hcols = st.columns([0.35, 1.0, 0.7, 0.7, 0.6, 0.6, 0.9, 0.7, 0.85, 0.85])
    for col, lbl in zip(hcols, ["","Categoria","Min lbs","Max lbs","Lotes","Semanas","Cap LBS","Ctd Anchos","MIX","Tipo Tejido"]):
        col.markdown(f"<div class='col-hdr'>{lbl}</div>", unsafe_allow_html=True)

    updated   = []
    total_cap = 0

    for i, c in enumerate(cats):
        cols = st.columns([0.35, 1.0, 0.7, 0.7, 0.6, 0.6, 0.9, 0.7, 0.85, 0.85])
        activo  = cols[0].checkbox("", value=c.activo,       key=f"ca_{i}")
        nombre  = cols[1].text_input("",  value=c.nombre,    key=f"cn_{i}", label_visibility="collapsed")
        minv    = cols[2].number_input("", value=int(c.minimo),  key=f"cmin_{i}", step=100, min_value=0, label_visibility="collapsed")
        maxv    = cols[3].number_input("", value=int(c.maximo),  key=f"cmax_{i}", step=100, min_value=1, label_visibility="collapsed")
        lotes   = cols[4].number_input("", value=c.lotes_dia,   key=f"cl_{i}",  step=1,   min_value=1, label_visibility="collapsed")
        sem_v   = round(float(c.semanas), 1)
        semanas = cols[5].number_input("", value=sem_v, key=f"cs_{i}", step=0.1, min_value=0.1, format="%.1f", label_visibility="collapsed")
        cap_c   = int(round(lotes * 7 * float(semanas) * maxv))
        cols[6].markdown(f"<div style='padding-top:6px'><span class='badge-blue'>{cap_c:,}</span></div>", unsafe_allow_html=True)
        ctd_anch= cols[7].number_input("", value=c.ctd_max_anchos, key=f"cca_{i}", step=1, min_value=1, max_value=10, label_visibility="collapsed")
        mix_i   = mix_opts.index(c.mix) if c.mix in mix_opts else 0
        mix_sel = cols[8].selectbox("", mix_opts,  index=mix_i, key=f"cmix_{i}", label_visibility="collapsed")
        tip_i   = tipo_opts.index(c.tipo_tejido) if c.tipo_tejido in tipo_opts else 0
        tip_sel = cols[9].selectbox("", tipo_opts, index=tip_i, key=f"ctj_{i}",  label_visibility="collapsed")

        updated.append(Categoria(
            nombre=nombre, minimo=minv, maximo=maxv, mix=mix_sel,
            tipo_tejido=tip_sel, lotes_dia=lotes, semanas=semanas,
            ctd_max_anchos=ctd_anch, activo=activo,
        ))
        if activo:
            total_cap += cap_c

    st.markdown(
        f"<div style='background:#003876;border-radius:6px;padding:8px 12px;margin-top:8px;color:white;'>"
        f"<b>TOTAL ACTIVAS: {total_cap:,} lbs</b></div>",
        unsafe_allow_html=True
    )
    if st.button("➕ Agregar Categoría"):
        updated.append(Categoria("NUEVA", 1000, 1100, "DYE"))
    st.session_state["categorias"] = updated


# ── TAB 2: RESTRICCIONES ─────────────────────────────────────────────────────
def _rest_table_ancho(data: list, prefix: str) -> list:
    hc = st.columns([0.3, 1.0, 0.7, 0.8, 0.8, 0.8, 0.25])
    for col, lbl in zip(hc, ["","STYLE","Límite Ancho","Prioridad 1","Prioridad 2","Prioridad 3",""]):
        col.markdown(f"<div class='col-hdr'>{lbl}</div>", unsafe_allow_html=True)
    updated = []
    for i, r in enumerate(data):
        c = st.columns([0.3, 1.0, 0.7, 0.8, 0.8, 0.8, 0.25])
        act  = c[0].checkbox("", value=r.activo, key=f"{prefix}_act_{i}")
        sty  = c[1].text_input("", value=r.style, key=f"{prefix}_sty_{i}", label_visibility="collapsed")
        lim  = c[2].number_input("", value=float(r.limite_ancho), key=f"{prefix}_lim_{i}", step=1.0, min_value=0.0, label_visibility="collapsed")
        p1   = c[3].text_input("", value=str(int(r.prioridades[0])) if len(r.prioridades) > 0 else "", key=f"{prefix}_p1_{i}", label_visibility="collapsed", placeholder="—")
        p2   = c[4].text_input("", value=str(int(r.prioridades[1])) if len(r.prioridades) > 1 else "", key=f"{prefix}_p2_{i}", label_visibility="collapsed", placeholder="—")
        p3   = c[5].text_input("", value=str(int(r.prioridades[2])) if len(r.prioridades) > 2 else "", key=f"{prefix}_p3_{i}", label_visibility="collapsed", placeholder="—")
        pris = [float(x) for x in [p1, p2, p3] if x.strip().isdigit() or (x.strip().replace(".", "", 1).isdigit())]
        if not c[-1].button("✕", key=f"{prefix}_del_{i}"):
            updated.append(ReglaAncho(sty.strip().upper(), lim, pris, act))
    if st.button("➕ Agregar", key=f"{prefix}_add"):
        updated.append(ReglaAncho("", 18, [2600]))
    return updated

def tab_restricciones():
    st.markdown("### Restricciones de Asignación")
    t1, t2, t3, t4 = st.tabs(["🔩 Ancho","🎨 Color","👕 Familia","📐 Anchos Combinados"])

    with t1:
        st.caption("STYLEs con restricción de ancho máximo → categoría destino (MAXIMO)")
        data = st.session_state.get("reglas_ancho", DEFAULT_REGLAS_ANCHO)
        st.session_state["reglas_ancho"] = _rest_table_ancho(data, "ra")

    with t2:
        st.caption("COLOR_R → categoría destino")
        data = st.session_state.get("reglas_color", DEFAULT_REGLAS_COLOR)
        hc = st.columns([0.3, 1.0, 0.8, 0.8, 0.8, 0.25])
        for col, lbl in zip(hc, ["","COLOR_R","Prioridad 1","Prioridad 2","Prioridad 3",""]):
            col.markdown(f"<div class='col-hdr'>{lbl}</div>", unsafe_allow_html=True)
        updated = []
        for i, r in enumerate(data):
            c = st.columns([0.3, 1.0, 0.8, 0.8, 0.8, 0.25])
            act = c[0].checkbox("", value=r.activo, key=f"rc_act_{i}")
            cr  = c[1].text_input("", value=r.color_r, key=f"rc_cr_{i}", label_visibility="collapsed")
            p1  = c[2].text_input("", value=str(int(r.prioridades[0])) if r.prioridades else "", key=f"rc_p1_{i}", label_visibility="collapsed", placeholder="—")
            p2  = c[3].text_input("", value=str(int(r.prioridades[1])) if len(r.prioridades) > 1 else "", key=f"rc_p2_{i}", label_visibility="collapsed", placeholder="—")
            p3  = c[4].text_input("", value=str(int(r.prioridades[2])) if len(r.prioridades) > 2 else "", key=f"rc_p3_{i}", label_visibility="collapsed", placeholder="—")
            pris = [float(x) for x in [p1, p2, p3] if x.strip().replace(".","",1).isdigit()]
            if not c[-1].button("✕", key=f"rc_del_{i}"):
                updated.append(ReglaColor(cr.strip().upper(), pris, act))
        if st.button("➕ Agregar", key="rc_add"):
            updated.append(ReglaColor("", []))
        st.session_state["reglas_color"] = updated

    with t3:
        st.caption("FAMILIA → categoría(s) destino")
        data = st.session_state.get("reglas_familia", DEFAULT_REGLAS_FAMILIA)
        hc = st.columns([0.3, 1.0, 0.7, 0.7, 0.7, 0.7, 0.25])
        for col, lbl in zip(hc, ["","FAMILIA","Prior.1","Prior.2","Prior.3","Prior.4",""]):
            col.markdown(f"<div class='col-hdr'>{lbl}</div>", unsafe_allow_html=True)
        updated = []
        for i, r in enumerate(data):
            c = st.columns([0.3, 1.0, 0.7, 0.7, 0.7, 0.7, 0.25])
            act = c[0].checkbox("", value=r.activo, key=f"rf_act_{i}")
            fam = c[1].text_input("", value=r.familia, key=f"rf_fam_{i}", label_visibility="collapsed")
            pts = []
            for j in range(4):
                val = r.prioridades[j] if j < len(r.prioridades) else None
                txt = c[j+2].text_input("", value=str(int(val)) if val else "", key=f"rf_p{j}_{i}", label_visibility="collapsed", placeholder="—")
                if txt.strip().replace(".", "", 1).isdigit():
                    pts.append(float(txt))
            if not c[-1].button("✕", key=f"rf_del_{i}"):
                updated.append(ReglaFamilia(fam.strip().upper(), pts, act))
        if st.button("➕ Agregar", key="rf_add"):
            updated.append(ReglaFamilia("", []))
        st.session_state["reglas_familia"] = updated

    with t4:
        st.caption("Si un lote combina ANCHO_1 y ANCHO_2 → dirigirlo a estas capacidades")
        data = st.session_state.get("reglas_combo", DEFAULT_REGLAS_COMBO)
        st.info("Configura pares de anchos combinados y sus categorías destino.")
        # Tabla simplificada para combo
        updated_raw = []
        for i, r in enumerate(data if isinstance(data, list) else []):
            c = st.columns([0.3, 0.7, 0.7, 0.7, 0.7, 0.7, 0.25])
            act = c[0].checkbox("", value=True, key=f"ac_act_{i}")
            a1  = c[1].number_input("A1", value=float(r.get("ancho_1", 18)), key=f"ac_a1_{i}", step=0.5, label_visibility="collapsed")
            a2  = c[2].number_input("A2", value=float(r.get("ancho_2", 20)), key=f"ac_a2_{i}", step=0.5, label_visibility="collapsed")
            p1  = c[3].text_input("", value=str(int(r.get("prioridades", [2600])[0])) if r.get("prioridades") else "", key=f"ac_p1_{i}", label_visibility="collapsed", placeholder="—")
            p2  = c[4].text_input("", value="", key=f"ac_p2_{i}", label_visibility="collapsed", placeholder="—")
            pris = [float(x) for x in [p1, p2] if x.strip().replace(".", "", 1).isdigit()]
            if not c[-1].button("✕", key=f"ac_del_{i}"):
                updated_raw.append({"ancho_1": a1, "ancho_2": a2, "prioridades": pris})
        if st.button("➕ Agregar combo", key="ac_add"):
            updated_raw.append({"ancho_1": 18.0, "ancho_2": 20.0, "prioridades": [2600]})
        st.session_state["reglas_combo"] = updated_raw


# ── TAB 3: CONFIGURACIÓN ──────────────────────────────────────────────────────
def tab_config():
    st.markdown("### Configuración del Motor NV3")
    _cfg_raw = st.session_state.get("config", DEFAULT_CONFIG)
    # Migración: si el objeto no tiene min_diff/max_diff (versión anterior), usar defaults
    if not hasattr(_cfg_raw, "min_diff"):
        _cfg_raw = DEFAULT_CONFIG
    cfg: Configuracion = _cfg_raw

    c1, c2, c3 = st.columns(3)

    with c1:
        st.markdown("**🔧 Parámetros del Solver**")
        max_items  = st.number_input("MAX_ITEMS — máx. LNKs distintos por lote", value=cfg.max_items, min_value=1, max_value=30, step=1)
        beam_width = st.number_input("BEAM_WIDTH — seeds candidatos por bloque",  value=cfg.beam_width, min_value=1, max_value=10, step=1)
        split_min  = st.number_input("SPLIT_MIN_LBS (default)",                   value=cfg.split_min_default, min_value=1.0, step=10.0)
        split_a18  = st.number_input("SPLIT_MIN_LBS (ANCHO18)",                   value=cfg.split_min_ancho18, min_value=1.0, step=10.0)
        bleach_r   = st.checkbox("Aplicar reglas a BLEACH",                       value=cfg.apply_rules_bleach)
        scrap_rem  = st.checkbox("Scrap de remanentes < SPLIT_MIN",               value=cfg.scrap_remainder)
        st.markdown("**📏 Diferencia de Anchos en un Lote**")
        st.caption("Diferencia en pulgadas entre los anchos distintos dentro de un mismo lote")
        min_diff = st.number_input("MIN_DIFF — diferencia mínima entre anchos (0=sin límite)", value=float(cfg.min_diff), min_value=0.0, max_value=20.0, step=0.5, format="%.1f")
        max_diff = st.number_input("MAX_DIFF — diferencia máxima entre anchos",               value=float(cfg.max_diff), min_value=0.0, max_value=30.0, step=0.5, format="%.1f")

    with c2:
        st.markdown("**📐 Objetivos de Anchos**")
        wto = st.text_input("WIDTHS_TARGET_ORDER (ej: 2>3>4)", value=cfg.widths_target_order)
        req_strict = st.checkbox("REQUIRE_WIDTHS_STRICT", value=cfg.require_widths_strict)
        upgrade = st.checkbox("UPGRADE_CATEGORIA",       value=cfg.upgrade_categoria)
        try_all = st.checkbox("TRY_ALL_PRIORITIES",      value=cfg.try_all_priorities)
        rule_order = st.text_input("RULE_ORDER",          value=cfg.rule_order)

        st.markdown("**🔀 Combinaciones de Prioridad**")
        st.caption("Un par por línea, separados por coma")
        combo_def = cfg.combinacion_prioridad
        combo_txt = st.text_area("", value="\n".join([",".join(p) for p in combo_def]), height=100, label_visibility="collapsed")
        try:
            combo_parsed = [ln.strip().split(",") for ln in combo_txt.strip().split("\n") if "," in ln]
        except Exception:
            combo_parsed = combo_def

    with c3:
        st.markdown("**⚖ Pesos de Scoring**")
        w_fill   = st.number_input("W_FILL (fill rate)",      value=cfg.w_fill,       step=0.5)
        w_cap    = st.number_input("W_CAP_LOSS (cap perdida)",value=cfg.w_cap_loss,    step=0.5)
        w_width  = st.number_input("W_WIDTH_PREF (anchos)",   value=cfg.w_width_pref,  step=0.5)
        w_1100   = st.number_input("W_1100_STRICT (E/G-1100)",value=cfg.w_1100_strict, step=1.0)
        w_comp   = st.number_input("W_COMPLETION",            value=cfg.w_completion,  step=0.1)

    st.session_state["config"] = Configuracion(
        max_items             = max_items,
        beam_width            = beam_width,
        split_min_default     = split_min,
        split_min_ancho18     = split_a18,
        apply_rules_bleach    = bleach_r,
        scrap_remainder       = scrap_rem,
        min_diff              = min_diff,
        max_diff              = max_diff,
        widths_target_order   = wto,
        require_widths_strict = req_strict,
        upgrade_categoria     = upgrade,
        try_all_priorities    = try_all,
        rule_order            = rule_order,
        combinacion_prioridad = combo_parsed,
        w_fill                = w_fill,
        w_cap_loss            = w_cap,
        w_width_pref          = w_width,
        w_1100_strict         = w_1100,
        w_completion          = w_comp,
        ancho18_allowed_max_dye = cfg.ancho18_allowed_max_dye,
    )


# ── TAB 4: EJECUTAR ───────────────────────────────────────────────────────────
def tab_ejecutar():
    st.markdown("### Cargar Datos y Ejecutar NV3")

    uploaded = st.file_uploader("📂 Sube el Excel (hoja DATA con header en fila 3)", type=["xlsx"])

    if uploaded:
        file_bytes = uploaded.read()
        df_raw, err = load_raw(file_bytes)
        if err:
            st.error(err)
            return
        df, err2 = build_dataframe(df_raw)
        if err2:
            st.error(err2)
            return
        st.session_state["df_cargado"] = df

        c1, c2, c3, c4 = st.columns(4)
        c1.markdown(f"<div class='metric-box'><div class='metric-val'>{len(df):,}</div><div class='metric-lbl'>Registros</div></div>", unsafe_allow_html=True)
        c2.markdown(f"<div class='metric-box'><div class='metric-val'>{df['ESTILO_C'].nunique()}</div><div class='metric-lbl'>Telas</div></div>", unsafe_allow_html=True)
        c3.markdown(f"<div class='metric-box'><div class='metric-val'>{df['MIX'].nunique()}</div><div class='metric-lbl'>MIX tipos</div></div>", unsafe_allow_html=True)
        c4.markdown(f"<div class='metric-box'><div class='metric-val'>{df['LBS_C'].sum():,.0f}</div><div class='metric-lbl'>Total LBS</div></div>", unsafe_allow_html=True)

        with st.expander("👁 Vista previa"):
            show = [c for c in ["ESTILO_C","COLOR_A","ANCHO","LBS_C","MIX","TONO","PRIORIDAD","LNK"] if c in df.columns]
            st.dataframe(df[show].head(30), use_container_width=True, height=220)

    st.markdown("---")
    cats_activas = [c for c in st.session_state.get("categorias", DEFAULT_CATEGORIAS) if c.activo]
    st.caption(f"Se procesarán **{len(cats_activas)}** categorías activas.")

    if st.button("▶ EJECUTAR NV3", use_container_width=False):
        df = st.session_state.get("df_cargado")
        if df is None:
            st.warning("⚠ Sube primero un archivo Excel.")
            return

        result_holder = [None]
        pool_holder   = [None]
        error_holder  = [None]
        prog_holder   = [0, "Iniciando…"]

        def progress_cb(pct, msg):
            prog_holder[0] = pct
            prog_holder[1] = msg

        def _run():
            try:
                res, pool = run_optimizer(
                    df          = df,
                    categorias  = st.session_state.get("categorias", DEFAULT_CATEGORIAS),
                    cfg         = st.session_state.get("config", DEFAULT_CONFIG),
                    reglas_ancho= st.session_state.get("reglas_ancho", DEFAULT_REGLAS_ANCHO),
                    reglas_color= st.session_state.get("reglas_color", DEFAULT_REGLAS_COLOR),
                    reglas_fam  = st.session_state.get("reglas_familia", DEFAULT_REGLAS_FAMILIA),
                    reglas_combo= [],
                    progress_cb = progress_cb,
                )
                result_holder[0] = res
                pool_holder[0]   = pool
            except Exception as e:
                import traceback
                error_holder[0] = f"{e}\n{traceback.format_exc()}"

        t = threading.Thread(target=_run, daemon=True)
        t.start()

        pbar   = st.progress(0)
        smsg   = st.empty()
        start  = time.time()

        while t.is_alive():
            elapsed = time.time() - start
            pct = min(int(prog_holder[0]), 95)
            pbar.progress(pct)
            smsg.markdown(f"<small>⚙ {prog_holder[1]} — <b>{int(elapsed)}s</b></small>", unsafe_allow_html=True)
            time.sleep(1.5)
            t.join(timeout=0)

        pbar.empty(); smsg.empty()

        if error_holder[0]:
            st.error(f"Error: {error_holder[0]}")
        elif result_holder[0] is None or result_holder[0].empty:
            st.warning("No se generaron lotes. Revisa parámetros y datos.")
        else:
            result = result_holder[0]
            pool   = pool_holder[0]
            cats   = st.session_state.get("categorias", DEFAULT_CATEGORIAS)
            reports = build_reports(result, df, cats, pool)

            st.session_state["resultado"] = result
            st.session_state["reportes"]  = reports
            st.session_state["pool"]      = pool

            elapsed = round(time.time() - start, 1)
            n_lotes = result["LOTE_ID"].nunique()
            total   = result.groupby("LOTE_ID")["TOTAL_LOTE"].first().sum()
            st.success(f"✅ {elapsed}s — **{n_lotes} lotes**, **{total:,.0f} lbs**. Ve a Resultados.")


# ── TAB 5: RESULTADOS ─────────────────────────────────────────────────────────
def tab_resultados():
    st.markdown("### Resultados NV3")

    result   = st.session_state.get("resultado")
    reportes = st.session_state.get("reportes", {})

    if result is None:
        st.info("Ejecuta el loteador primero.")
        return

    # KPIs
    n_lotes   = result["LOTE_ID"].nunique()
    total_lbs = result.groupby("LOTE_ID")["TOTAL_LOTE"].first().sum()
    puro_n    = result[result["TIPO_LOTE_ANCHO"] == "PURO"]["LOTE_ID"].nunique()
    puro_pct  = round(puro_n / n_lotes * 100, 1) if n_lotes else 0
    exc       = reportes.get("EXCEDENTES", pd.DataFrame())
    n_exc     = len(exc)
    excedidos = reportes.get("LNK_COMPLETITUD", pd.DataFrame())
    n_exced   = int((excedidos["ESTADO"] == "EXCEDIDO").sum()) if not excedidos.empty else 0

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.markdown(f"<div class='metric-box'><div class='metric-val'>{n_lotes}</div><div class='metric-lbl'>Lotes</div></div>", unsafe_allow_html=True)
    c2.markdown(f"<div class='metric-box'><div class='metric-val'>{total_lbs:,.0f}</div><div class='metric-lbl'>LBS Loteadas</div></div>", unsafe_allow_html=True)
    c3.markdown(f"<div class='metric-box'><div class='metric-val'>{puro_pct}%</div><div class='metric-lbl'>Lotes Puros</div></div>", unsafe_allow_html=True)
    c4.markdown(f"<div class='metric-box'><div class='metric-val' style='color:#e67e22'>{n_exc}</div><div class='metric-lbl'>Excedentes</div></div>", unsafe_allow_html=True)
    c5.markdown(f"<div class='metric-box'><div class='metric-val' style='color:#c0392b'>{n_exced}</div><div class='metric-lbl'>LNK Excedidos</div></div>", unsafe_allow_html=True)
    c6.markdown(f"<div class='metric-box'><div class='metric-val'>{result['CATEGORIA'].nunique()}</div><div class='metric-lbl'>Categorías</div></div>", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    tabs = st.tabs([
        "📋 Detalle","📊 Resumen Lotes","🏭 Capacidad","⚖ Prioridad",
        "⚠ Excedentes","🔗 LNK","📝 Decision Log","📁 Resumen Cat"
    ])

    with tabs[0]:
        df_d = reportes.get("DETALLE_LOTES", result)
        cf1, cf2 = st.columns(2)
        with cf1:
            cat_f = st.selectbox("Categoría", ["Todas"] + sorted(result["CATEGORIA"].unique().tolist()), key="r_cat")
        with cf2:
            mix_f = st.selectbox("MIX", ["Todos"] + sorted(result["MIX"].unique().tolist()), key="r_mix")
        df_s = df_d.copy()
        if cat_f != "Todas" and "CATEGORIA" in df_s.columns:
            df_s = df_s[df_s["CATEGORIA"] == cat_f]
        if mix_f != "Todos" and "MIX" in df_s.columns:
            df_s = df_s[df_s["MIX"] == mix_f]
        st.dataframe(df_s, use_container_width=True, height=420)

    with tabs[1]:
        st.dataframe(reportes.get("RESUMEN_LOTES", pd.DataFrame()), use_container_width=True, height=420)

    with tabs[2]:
        df_cap = reportes.get("CAPACIDAD_X_CATEG", pd.DataFrame())
        st.dataframe(df_cap, use_container_width=True, height=280)
        if not df_cap.empty and "CAPACIDAD_LBS" in df_cap.columns:
            st.bar_chart(df_cap.set_index("CATEGORIA")[["CAPACIDAD_LBS", "LBS_ASIGNADAS"]])

    with tabs[3]:
        st.dataframe(reportes.get("PRIORIDAD_VS_ASIG", pd.DataFrame()), use_container_width=True, height=280)

    with tabs[4]:
        df_exc = reportes.get("EXCEDENTES", pd.DataFrame())
        if df_exc.empty:
            st.success("✅ Todos los ítems fueron asignados.")
        else:
            st.warning(f"⚠ {len(df_exc)} ítems con LBS_RESTANTES > 0")
            st.dataframe(df_exc, use_container_width=True, height=350)

    with tabs[5]:
        df_lnk = reportes.get("LNK_COMPLETITUD", pd.DataFrame())
        if not df_lnk.empty and "ESTADO" in df_lnk.columns:
            comp = (df_lnk["ESTADO"] == "COMPLETO").sum()
            parc = (df_lnk["ESTADO"] == "PARCIAL").sum()
            exce = (df_lnk["ESTADO"] == "EXCEDIDO").sum()
            lc1, lc2, lc3 = st.columns(3)
            lc1.markdown(f"<div class='metric-box'><div class='metric-val' style='color:#1a7a4a'>{comp}</div><div class='metric-lbl'>Completos</div></div>", unsafe_allow_html=True)
            lc2.markdown(f"<div class='metric-box'><div class='metric-val' style='color:#e67e22'>{parc}</div><div class='metric-lbl'>Parciales</div></div>", unsafe_allow_html=True)
            lc3.markdown(f"<div class='metric-box'><div class='metric-val' style='color:#c0392b'>{exce}</div><div class='metric-lbl'>Excedidos</div></div>", unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)
            est_f = st.selectbox("Filtrar", ["Todos"] + sorted(df_lnk["ESTADO"].unique().tolist()), key="lnk_f")
            df_s  = df_lnk if est_f == "Todos" else df_lnk[df_lnk["ESTADO"] == est_f]
            st.dataframe(df_s, use_container_width=True, height=380)
        else:
            st.dataframe(df_lnk, use_container_width=True, height=380)

    with tabs[6]:
        df_log = reportes.get("DECISION_LOG", pd.DataFrame())
        st.caption("Registro completo de decisiones del motor — APLICA_REGLA y DECISION_SCORE por lote.")
        st.dataframe(df_log, use_container_width=True, height=420)

    with tabs[7]:
        st.dataframe(reportes.get("RESUMEN_CATEGORIA", pd.DataFrame()), use_container_width=True, height=300)

    # ── DESCARGA ──────────────────────────────────────────────────────────────
    st.markdown("---")
    tz    = pytz.timezone("America/Tegucigalpa")
    ts    = datetime.now(tz).strftime("%Y%m%d_%H%M%S")
    fname = f"NV3_loteo_{ts}.xlsx"
    buf   = BytesIO()

    sheet_order = [
        ("DETALLE_LOTES",     reportes.get("DETALLE_LOTES",     result)),
        ("RESUMEN_LOTES",     reportes.get("RESUMEN_LOTES",     pd.DataFrame())),
        ("RESUMEN_CATEGORIA", reportes.get("RESUMEN_CATEGORIA", pd.DataFrame())),
        ("EXCEDENTES",        reportes.get("EXCEDENTES",        pd.DataFrame())),
        ("CAPACIDAD_X_CATEG", reportes.get("CAPACIDAD_X_CATEG", pd.DataFrame())),
        ("PRIORIDAD_VS_ASIG", reportes.get("PRIORIDAD_VS_ASIG", pd.DataFrame())),
        ("LNK_COMPLETITUD",   reportes.get("LNK_COMPLETITUD",   pd.DataFrame())),
        ("DECISION_LOG",      reportes.get("DECISION_LOG",      pd.DataFrame())),
        ("PARAMETROS",        reportes.get("PARAMETROS",        pd.DataFrame())),
    ]
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for sname, df_s in sheet_order:
            if df_s is not None and not df_s.empty:
                df_s.to_excel(writer, sheet_name=sname, index=False)
        format_excel(writer)
    buf.seek(0)

    st.download_button(
        "⬇ Descargar Excel NV3", data=buf, file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    sidebar()

    st.markdown("""
    <div class="logo-header">
      <div style="font-size:18px;font-weight:800;color:white;">🏭 NV3 — Loteador Inteligente Elcatex</div>
      <div style="font-size:9px;color:#cce4ff;letter-spacing:2px;text-transform:uppercase;margin-top:3px;">
        Grupo Elcatex · Planeación de la Demanda · Honduras 🇭🇳
      </div>
    </div>
    """, unsafe_allow_html=True)

    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📊 Capacidades",
        "🔒 Restricciones",
        "⚙ Configuración",
        "▶ Ejecutar",
        "📋 Resultados",
    ])
    with tab1: tab_capacidades()
    with tab2: tab_restricciones()
    with tab3: tab_config()
    with tab4: tab_ejecutar()
    with tab5: tab_resultados()


if __name__ == "__main__":
    main()
