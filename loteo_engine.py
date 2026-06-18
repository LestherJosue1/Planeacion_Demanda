# ============================================
# loteo_engine.py
# Motor puro (sin Streamlit / sin Colab) del algoritmo de loteo de tintorería NV2.
# Adaptado del script original "ANALYSYS_DATA_2026_Pref_anchos_.txt".
#
# CAMBIOS RESPECTO AL SCRIPT ORIGINAL (acordados con el usuario):
#  - Se ELIMINÓ el modo HUMANO % (OVERSHOOT_ENABLE / UNDERSHOOT_ENABLE / choose_take_humano).
#    Siempre se usa choose_take() (reparto exacto, sin overshoot/undershoot).
#  - SPLIT_MIN_LBS_DEFAULT default = 500 (antes 100).
#  - MAX_WIDTHS ahora es POR CATEGORÍA de tintorería (reemplaza el MAX_WIDTHS global único).
#  - COMBINACION_PRIORIDAD: matriz fija de bloques que se pueden mezclar:
#        VENCIDOS-VENCIDOS, VENCIDOS-AHEAD, AHEAD-AHEAD, AHEAD-AHEAD2, OTROS-AHEAD2
#    (editable desde la UI vía checkboxes, no se parsea texto libre en producción).
#  - TIPO_TEJIDO: nuevo parámetro de scoring. Si la categoría del lote es A-4000 o B-3300,
#    el tejido del SKU semilla (columna TIPO_TEJIDO) es FLEECE, y la FAMILIA de ese SKU NO
#    tiene RESTRICCION_FAMILIA activa, se suma un bono W_TIPO_TEJIDO_FLEECE al score del lote.
#  - %CARGA: nueva columna por fila en DATA (decimal, ej 0.8, 1.0). Reduce el MAXIMO efectivo
#    de capacidad de la categoría para ese lote (ej. 1100*0.80=880) sin cambiar la categoría
#    asignada (sigue siendo E-1100/G-1100). Se aplica usando el %CARGA del SKU semilla del lote.
# ============================================

import pandas as pd
import numpy as np
import re
from itertools import permutations

# ---------------------------- Utils ----------------------------
def norm_str(x):
    if pd.isna(x):
        return ""
    return str(x).strip()

def up(x):
    return norm_str(x).upper()

def clean_cols(cols):
    out = []
    for c in cols:
        c = "" if c is None else str(c)
        c = c.replace("\n", " ").replace("\r", " ")
        c = re.sub(r"\s+", " ", c).strip()
        out.append(c)
    return out

# ---------------------------- Readers ----------------------------
def find_header_row(xlsm_path, sheet_name, required_cols, search_rows=80):
    preview = pd.read_excel(xlsm_path, sheet_name=sheet_name, engine="openpyxl", header=None, nrows=search_rows)
    req = set(required_cols)
    for r in range(min(search_rows, len(preview))):
        row_vals = [norm_str(v) for v in preview.iloc[r].tolist()]
        row_set = set([v for v in row_vals if v])
        if req.issubset(row_set):
            return r
    return 0

def read_sheet_autoheader(xlsm_path, sheet_name, required_cols=None, default_header=0):
    hdr = find_header_row(xlsm_path, sheet_name, required_cols) if required_cols else default_header
    df = pd.read_excel(xlsm_path, sheet_name=sheet_name, engine="openpyxl", header=hdr)
    df.columns = clean_cols(df.columns)
    return df, hdr

# ---------------------------- Blocks & Widths ----------------------------
def prioridad_bloque(prio_text: str) -> str:
    p = (prio_text or "").upper()
    if "PAST DUE" in p or "DUE" in p or "VENC" in p:
        return "VENCIDOS"
    if "AHEAD2" in p:
        return "AHEAD2"
    if "AHEAD" in p:
        return "AHEAD"
    return "OTROS"

def can_mix_blocks(b1, b2, allowed_pairs):
    if b1 == b2:
        return True
    return (b1, b2) in allowed_pairs

def valid_width_group(widths, min_diff, max_diff, max_widths):
    w = [float(x) for x in widths if x is not None and not pd.isna(x) and float(x) != 0.0]
    uw = sorted(set(w))
    if len(uw) <= 1:
        return True
    if len(uw) > int(max_widths):
        return False
    for i in range(len(uw)):
        for j in range(i + 1, len(uw)):
            d = abs(uw[j] - uw[i])
            if d < min_diff or d > max_diff:
                return False
    return True

def get_row_widths(work, idx):
    widths = []
    for c in ["ANCHO.F.C", "ANCHO.F.M"]:
        if c in work.columns:
            v = work.at[idx, c]
            if pd.notna(v) and float(v) != 0.0:
                widths.append(float(v))
    return widths

# ---------------------------- Split chooser ----------------------------
def choose_take(rest, remaining, split_min_lbs, allow_scrap_residue=False):
    try:
        split_min_lbs = float(split_min_lbs)
    except Exception:
        split_min_lbs = 0.0
    if rest <= 0 or remaining <= 0:
        return 0.0
    if rest <= remaining + 1e-9:
        return float(rest)
    take = float(remaining)
    if take + 1e-9 < split_min_lbs:
        return 0.0
    residue = float(rest) - take
    if residue > 1e-9 and residue + 1e-9 < split_min_lbs:
        if not allow_scrap_residue:
            return 0.0
        return take
    return take

# ---------------------------- Ranges builder ----------------------------
def build_ranges(df_cap):
    ranges = []
    for _, r in df_cap.iterrows():
        ranges.append({
            "CATEGORIA": norm_str(r["CATEGORIA"]),
            "MINIMO": float(r["MINIMO"]),
            "MAXIMO": float(r["MAXIMO"]),
            "CAPACIDAD": float(r["CAPACIDAD"]),
            "MIX": up(r["MIX"]),
            "RANGO_ID": f"CAP_{norm_str(r['CATEGORIA'])}_{up(r['MIX'])}_{float(r['MAXIMO']):.0f}"
        })
    return sorted(ranges, key=lambda x: x["MAXIMO"], reverse=True)

# ---------------------------- Generación de órdenes posibles para ORDEN_REGLAS ----------------------------
def all_rule_order_options():
    """Devuelve todas las combinaciones posibles (permutaciones) de las 4 reglas
    + DEFAULT siempre al final, para exponer en la UI como selector de escenarios."""
    base_rules = ["ANCHO18", "COMBO_ANCHOS", "COLOR_R", "FAMILIA"]
    opts = []
    for perm in permutations(base_rules):
        opts.append(">".join(perm) + ">DEFAULT")
    return opts

# ---------------------------- Load inputs (DATA + params ya parametrizados) ----------------------------
REQUIRED_DATA_COLS = ["LNK", "TELA.CUERPO", "COLOR", "PRIORIDAD", "ANCHO.F.C", "ANCHO.F.M", "TOTAL", "MIX", "CONSUMO_C"]

def load_data_sheet(xlsm_path):
    """Lee únicamente la hoja DATA con autodetección de fila de encabezado."""
    df_data, hdr_row = read_sheet_autoheader(xlsm_path, "DATA", required_cols=REQUIRED_DATA_COLS, default_header=0)
    miss = [c for c in REQUIRED_DATA_COLS if c not in df_data.columns]
    if miss:
        raise ValueError(f"DATA: faltan columnas obligatorias {miss}. Header detectado en fila {hdr_row+1}.")

    for c in ["ANCHO.F.C", "ANCHO.F.M", "TOTAL", "CONSUMO_C"]:
        df_data[c] = pd.to_numeric(df_data[c], errors="coerce").fillna(0.0)
    for c in ["LNK", "TELA.CUERPO", "COLOR", "PRIORIDAD", "MIX"]:
        df_data[c] = df_data[c].apply(norm_str)
    df_data["MIX"] = df_data["MIX"].apply(up)

    for opt_col, default in [("FAMILIA", ""), ("COLOR_R", ""), ("STYLE", "")]:
        if opt_col not in df_data.columns:
            df_data[opt_col] = default
        else:
            df_data[opt_col] = df_data[opt_col].apply(up)

    if "TONO" in df_data.columns:
        df_data["TONO"] = df_data["TONO"].apply(up)

    # TIPO_TEJIDO (nueva columna): FLEECE / JERSEY / OTRO
    if "TIPO_TEJIDO" not in df_data.columns:
        df_data["TIPO_TEJIDO"] = ""
    else:
        df_data["TIPO_TEJIDO"] = df_data["TIPO_TEJIDO"].apply(up)

    # %CARGA (nueva columna): decimal por fila, ej 0.7, 0.8, 1.0
    carga_col = None
    for cand in ["%CARGA", "PCT_CARGA", "PORCENTAJE_CARGA", "% CARGA"]:
        if cand in df_data.columns:
            carga_col = cand
            break
    if carga_col is None:
        df_data["PCT_CARGA"] = 1.0
    else:
        df_data["PCT_CARGA"] = pd.to_numeric(df_data[carga_col], errors="coerce").fillna(1.0)
        df_data.loc[(df_data["PCT_CARGA"] <= 0) | (df_data["PCT_CARGA"] > 1.0), "PCT_CARGA"] = 1.0

    return df_data, hdr_row

def build_cap_dataframe(cap_rows):
    """cap_rows: lista de dicts {CATEGORIA, MINIMO, MAXIMO, CAPACIDAD, MIX} (viene del data_editor de la UI)."""
    df_cap = pd.DataFrame(cap_rows)
    df_cap["CATEGORIA"] = df_cap["CATEGORIA"].apply(norm_str)
    df_cap["MIX"] = df_cap["MIX"].apply(up)
    for c in ["MINIMO", "MAXIMO", "CAPACIDAD"]:
        df_cap[c] = pd.to_numeric(df_cap[c], errors="coerce")
    if df_cap[["MINIMO", "MAXIMO", "CAPACIDAD"]].isna().any().any():
        raise ValueError("CAPACIDAD TINTORERIA: hay valores MINIMO/MAXIMO/CAPACIDAD inválidos o vacíos.")
    return df_cap

# ---------------------------- Priority helpers ----------------------------
def order_priorities(pris, params):
    pris = [float(x) for x in pris if x is not None]
    po_text = norm_str(params.get("PRIORITY_ORDER", ""))
    if po_text:
        plan = [p.strip() for p in po_text.split(">") if p.strip()]
        rank = {}
        for i, v in enumerate(plan):
            if re.match(r"^\d+(\.\d+)?$", v):
                rank[float(v)] = i
        return sorted(pris, key=lambda x: (rank.get(float(x), 10_000), float(x)))
    return sorted(pris)

def order_by_priorities(base_ranges, prioridades):
    used = set()
    out = []
    for cap in prioridades:
        match = [r for r in base_ranges if abs(float(r["MAXIMO"]) - float(cap)) < 1e-6]
        for r in match:
            if id(r) not in used:
                out.append(r); used.add(id(r))
    for r in base_ranges:
        if id(r) not in used:
            out.append(r)
    return out

# ---------------------------- Reorder rules (ANCHO18 / COMBO_ANCHOS / COLOR_R / FAMILIA) ----------------------------
def reorder_ranges_for_seed(ranges_mix, mixv, work, seed_idx, params):
    base = list(ranges_mix)
    rule_info = {
        "regla_aplicada": "NONE",
        "prioridades": [],
        "match_combo": False,
        "limite_ancho_style": None,
        "origen_prioridad": "MIX",
        "combo_target_width": None,
    }
    if up(mixv) not in ("DYE",) and int(params.get("APPLY_RULES_BLEACH", 0)) != 1:
        return base, rule_info

    fam = up(work.at[seed_idx, "FAMILIA"]) if "FAMILIA" in work.columns else ""
    color_r = up(work.at[seed_idx, "COLOR_R"]) if "COLOR_R" in work.columns else ""
    style = up(work.at[seed_idx, "STYLE"]) if "STYLE" in work.columns else ""

    def f2(x):
        try:
            return float(x)
        except Exception:
            return 0.0

    ancho_c = f2(work.at[seed_idx, "ANCHO.F.C"]) if "ANCHO.F.C" in work.columns else 0.0
    ancho_m = f2(work.at[seed_idx, "ANCHO.F.M"]) if "ANCHO.F.M" in work.columns else 0.0

    restr_fam = params.get("RESTRICCIONES_FAMILIA", {})
    restr_color = params.get("RESTRICCIONES_COLOR", {})
    restr_ancho = params.get("RESTRICCIONES_ANCHO", {})
    reglas_combo = params.get("REGLAS_ANCHOS_COMBINADOS", [])
    rule_order_cfg = norm_str(params.get("RULE_ORDER", ""))
    rule_order = [x.strip().upper() for x in rule_order_cfg.split(">") if x.strip()] or \
        ["ANCHO18", "COMBO_ANCHOS", "COLOR_R", "FAMILIA", "DEFAULT"]

    def ancho_activo_leq_lim(ac, am, lim):
        vals = []
        try:
            if ac is not None and not pd.isna(ac) and float(ac) > 0:
                vals.append(float(ac))
        except Exception:
            pass
        try:
            if am is not None and not pd.isna(am) and float(am) > 0:
                vals.append(float(am))
        except Exception:
            pass
        return (len(vals) > 0) and (min(vals) <= float(lim))

    def try_ancho18():
        if style in restr_ancho:
            lim = restr_ancho[style].get("limite", None)
            prioridades = order_priorities(restr_ancho[style].get("prioridades", []), params)
            if lim is not None and ancho_activo_leq_lim(ancho_c, ancho_m, lim) and len(prioridades) > 0:
                rule_info.update({
                    "regla_aplicada": "ANCHO18",
                    "prioridades": list(prioridades),
                    "limite_ancho_style": lim,
                    "origen_prioridad": "STYLE",
                })
                return order_by_priorities(base, prioridades)
        return None

    def try_combo():
        for regla in reglas_combo:
            a1, a2 = regla["a1"], regla["a2"]
            prioridades = order_priorities(regla["prioridades"], params)
            seed_matches = (abs(ancho_c - a1) < 1e-6 or abs(ancho_m - a1) < 1e-6 or
                            abs(ancho_c - a2) < 1e-6 or abs(ancho_m - a2) < 1e-6)
            if not seed_matches:
                continue
            objetivo = a2 if (abs(ancho_c - a1) < 1e-6 or abs(ancho_m - a1) < 1e-6) else a1
            existe_otro = False
            for idx in work.index:
                if idx == seed_idx:
                    continue
                if float(work.at[idx, "LBS_RESTANTES"]) <= 0:
                    continue
                ac = f2(work.at[idx, "ANCHO.F.C"]); am = f2(work.at[idx, "ANCHO.F.M"])
                if abs(ac - objetivo) < 1e-6 or abs(am - objetivo) < 1e-6:
                    existe_otro = True; break
            if existe_otro and len(prioridades) > 0:
                rule_info.update({
                    "regla_aplicada": "COMBO_ANCHOS",
                    "prioridades": list(prioridades),
                    "match_combo": True,
                    "origen_prioridad": "COMBO",
                    "combo_target_width": float(objetivo),
                })
                return order_by_priorities(base, prioridades)
        return None

    def try_color_r():
        if color_r in restr_color and restr_color[color_r]:
            p = float(restr_color[color_r])
            rule_info.update({
                "regla_aplicada": "COLOR_R",
                "prioridades": [p],
                "origen_prioridad": "COLOR"
            })
            return order_by_priorities(base, [p])
        return None

    def try_familia():
        if fam in restr_fam and len(restr_fam[fam]) > 0:
            prioridades = order_priorities(restr_fam[fam], params)
            rule_info.update({
                "regla_aplicada": "FAMILIA",
                "prioridades": list(prioridades),
                "origen_prioridad": "FAMILIA"
            })
            return order_by_priorities(base, prioridades)
        return None

    for token in rule_order:
        out = None
        if token == "ANCHO18":
            out = try_ancho18()
        elif token == "COMBO_ANCHOS":
            out = try_combo()
        elif token == "COLOR_R":
            out = try_color_r()
        elif token == "FAMILIA":
            out = try_familia()
        elif token == "DEFAULT":
            out = base
        if out is not None:
            return out, rule_info
    return base, rule_info

# ---------------------------- Priority matching for upgrades ----------------------------
def ranges_matching_priority(pri, ranges_try, tol=1e-6, allow_nearest_higher=True):
    pri = float(pri)
    exact = [r for r in ranges_try if abs(float(r["MAXIMO"]) - pri) <= tol]
    if exact:
        return exact
    if not allow_nearest_higher:
        return []
    higher = [r for r in ranges_try if float(r["MAXIMO"]) >= pri - tol]
    if higher:
        higher = sorted(higher, key=lambda r: (float(r["MAXIMO"]) - pri, -float(r["MAXIMO"])))
        return [higher[0]]
    return []

# ---------------------------- Scoring ----------------------------
def score_lote(lote_dict, resumen_rows, params, categoria=None, seed_row=None):
    if lote_dict is None:
        return -1e30
    W_FILL = params.get("W_FILL", 5.0)
    W_CAP_LOSS = params.get("W_CAP_LOSS", 3.0)
    W_WIDTH_PREF = params.get("W_WIDTH_PREF", 2.0)
    W_1100_STRICT = params.get("W_1100_WIDTHS_STRICT", 10.0)
    pref_list = params.get("WIDTH_PREF_LIST", [2, 3, 1, 4, 5, 6])

    total = float(lote_dict.get("TOTAL_LOTE", 0.0))
    maximo = float(lote_dict.get("MAXIMO", 1.0))
    fill = total / maximo if maximo > 1e-9 else 0.0
    cap_loss = (maximo - total)

    anchos = set()
    for r in resumen_rows:
        for w in r.get("ANCHOS_ROW", []):
            if w is not None:
                anchos.add(float(w))
    widths_unique = len(anchos)

    try:
        rank = pref_list.index(widths_unique)
    except ValueError:
        rank = len(pref_list) + abs(widths_unique - pref_list[-1])
    width_pref_score = -float(rank)

    score = (W_FILL * fill) + (-W_CAP_LOSS * cap_loss) + (W_WIDTH_PREF * width_pref_score)
    if abs(maximo - 1100.0) < 1e-6:
        score -= W_1100_STRICT * max(0, widths_unique - 1)

    # --- NUEVO: TIPO_TEJIDO (preferencia FLEECE en categorías grandes) ---
    if int(params.get("TIPO_TEJIDO_ENABLE", 0)) == 1 and categoria is not None and seed_row is not None:
        cats_fleece = set(params.get("TIPO_TEJIDO_CATEGORIAS", ["A-4000", "B-3300"]))
        if categoria in cats_fleece:
            tejido = up(seed_row.get("TIPO_TEJIDO", ""))
            familia = up(seed_row.get("FAMILIA", ""))
            restr_fam = params.get("RESTRICCIONES_FAMILIA", {})
            familia_tiene_restriccion = familia in restr_fam and len(restr_fam.get(familia, [])) > 0
            if tejido == "FLEECE" and not familia_tiene_restriccion:
                score += float(params.get("W_TIPO_TEJIDO_FLEECE", 4.0))
    return score

# ---------------------------- Filtro por objetivo de # de anchos ----------------------------
def filter_ranges_for_width_target(ranges_try, mixv, width_target, params):
    mixu = str(mixv).strip().upper()
    allowed = None
    if width_target == 3:
        allowed_map = params.get("ALLOWED_MAXIMO_FOR_3_WIDTHS", {})
        allowed = allowed_map.get(mixu, set())
    elif width_target == 4:
        allowed_map = params.get("ALLOWED_MAXIMO_FOR_4_WIDTHS", {})
        allowed = allowed_map.get(mixu, set())
    if allowed and len(allowed) > 0:
        return [r for r in ranges_try if float(r["MAXIMO"]) in allowed]
    return list(ranges_try)

# ---------------------------- Intento de lote ----------------------------
def intentar_lote_para_rango(work, seed_idx, rango, capacity_used, params, rule_info,
                              require_two_widths=False, split_min_lbs=None,
                              min_unique_widths=None, max_unique_widths=None):
    min_diff = params["MIN_DIFF"]
    max_diff = params["MAX_DIFF"]
    max_sku = params["MAX_SKU"]
    allowed_pairs = params["MIX_ALLOWED"]

    # MAX_WIDTHS por categoría (reemplaza global)
    max_widths_by_cat = params.get("MAX_WIDTHS_BY_CAT", {})
    max_widths = max_widths_by_cat.get(rango["CATEGORIA"], params.get("MAX_WIDTHS_DEFAULT", 4))

    rid = rango["RANGO_ID"]
    cap_total = float(rango["CAPACIDAD"])
    cap_used = float(capacity_used.get(rid, 0.0))
    cap_left_global = max(0.0, cap_total - cap_used)
    if cap_left_global <= 0:
        return None

    # %CARGA: el MAXIMO efectivo del lote se reduce según el % de carga del SKU semilla
    pct_carga_seed = 1.0
    if "PCT_CARGA" in work.columns:
        try:
            pct_carga_seed = float(work.at[seed_idx, "PCT_CARGA"])
            if pct_carga_seed <= 0 or pct_carga_seed > 1.0:
                pct_carga_seed = 1.0
        except Exception:
            pct_carga_seed = 1.0

    max_allowed = min(float(rango["MAXIMO"]) * pct_carga_seed, cap_left_global)

    if float(work.at[seed_idx, "LBS_RESTANTES"]) <= 0:
        return None

    try:
        split_min_lbs = float(split_min_lbs if split_min_lbs is not None else params.get("SPLIT_MIN_LBS_DEFAULT", 500.0))
    except Exception:
        split_min_lbs = float(params.get("SPLIT_MIN_LBS_DEFAULT", 500.0))
    allow_scrap_residue = int(params.get("SCRAP_REMAINDER_BELOW_SPLIT_MIN", 1)) == 1

    lote_rows = []
    lote_lbs = 0.0
    lote_lnks = set()
    lote_blocks = []
    lote_widths = []

    def can_add_row(idx, lbs_to_add):
        if lbs_to_add <= 0:
            return False
        if "TONO" in work.columns:
            seed_tono = up(work.at[seed_idx, "TONO"]) if not pd.isna(work.at[seed_idx, "TONO"]) else ""
            row_tono = up(work.at[idx, "TONO"]) if not pd.isna(work.at[idx, "TONO"]) else ""
            if seed_tono != row_tono:
                return False

        lnk = work.at[idx, "LNK"]
        new_lnks = set(lote_lnks); new_lnks.add(lnk)
        if len(new_lnks) > max_sku:
            return False

        b = work.at[idx, "BLOQUE"]
        for existing_b in lote_blocks:
            if not can_mix_blocks(existing_b, b, allowed_pairs):
                return False

        widths_candidate = list(lote_widths) + get_row_widths(work, idx)
        if not valid_width_group(widths_candidate, min_diff, max_diff, max_widths):
            return False

        if max_unique_widths is not None:
            uwc = sorted(set([float(w) for w in widths_candidate if w is not None and not pd.isna(w) and float(w) != 0.0]))
            if len(uwc) > int(max_unique_widths):
                return False

        if lote_lbs + lbs_to_add > max_allowed + 1e-9:
            return False
        return True

    seed_rest = float(work.at[seed_idx, "LBS_RESTANTES"])
    remaining = max_allowed - lote_lbs
    take = choose_take(seed_rest, remaining, split_min_lbs, allow_scrap_residue=allow_scrap_residue)

    if take <= 0 or not can_add_row(seed_idx, take):
        return None

    lote_rows.append((seed_idx, take, 0.0, 0.0))
    lote_lbs += take
    lote_lnks.add(work.at[seed_idx, "LNK"])
    lote_blocks.append(work.at[seed_idx, "BLOQUE"])
    lote_widths += get_row_widths(work, seed_idx)

    combo_target = rule_info.get("combo_target_width", None) if rule_info else None

    while True:
        remaining = max_allowed - lote_lbs
        if remaining <= 1e-6:
            break
        best = None
        best_take = 0.0
        best_score = -1e30

        for idx in work.index:
            rest = float(work.at[idx, "LBS_RESTANTES"])
            if rest <= 0:
                continue
            if any(i == idx for i, *_ in lote_rows):
                continue
            take = choose_take(rest, remaining, split_min_lbs, allow_scrap_residue=allow_scrap_residue)
            if take <= 0:
                continue
            if not can_add_row(idx, take):
                continue

            new_total = lote_lbs + take
            widths_now = set([float(w) for w in lote_widths if w is not None and not pd.isna(w) and float(w) != 0.0])
            widths_add = set([float(w) for w in get_row_widths(work, idx) if w is not None and not pd.isna(w) and float(w) != 0.0])
            new_widths = widths_now.union(widths_add)
            adds_new_width = 1 if len(new_widths) > len(widths_now) else 0

            has_target = 0
            if combo_target is not None:
                for w in widths_add:
                    if abs(float(w) - float(combo_target)) < 1e-6:
                        has_target = 1
                        break

            score = new_total + has_target * 1e-3 + adds_new_width * 1e-4
            if score > best_score:
                best_score = score
                best = idx
                best_take = take

        if best is None:
            break

        lote_rows.append((best, best_take, 0.0, 0.0))
        lote_lbs += best_take
        lote_lnks.add(work.at[best, "LNK"])
        lote_blocks.append(work.at[best, "BLOQUE"])
        lote_widths += get_row_widths(work, best)

    if lote_lbs + 1e-9 < float(rango["MINIMO"]) * pct_carga_seed:
        return None

    if min_unique_widths is not None:
        min_required = int(min_unique_widths)
    elif require_two_widths:
        min_required = 2
    else:
        min_required = None

    if min_required is not None:
        uw = sorted(set([float(w) for w in lote_widths if w is not None and not pd.isna(w) and float(w) != 0.0]))
        if len(uw) < int(min_required):
            return None

    if max_unique_widths is not None:
        uw = sorted(set([float(w) for w in lote_widths if w is not None and not pd.isna(w) and float(w) != 0.0]))
        if len(uw) > int(max_unique_widths):
            return None

    return {
        "RANGO_ID": rango["RANGO_ID"],
        "CATEGORIA": rango["CATEGORIA"],
        "MIX": rango["MIX"],
        "MINIMO": float(rango["MINIMO"]),
        "MAXIMO": float(rango["MAXIMO"]),
        "TOTAL_LOTE": float(lote_lbs),
        "ROWS": lote_rows,
        "REQUIERE_2_ANCHOS": bool(require_two_widths),
        "PCT_CARGA_USADO": pct_carga_seed,
    }

# ---------------------------- Loteo principal ----------------------------
def run_loteo(df_data, df_cap, params, progress_cb=None):
    """progress_cb(fraction:float, msg:str) opcional para barra de progreso en Streamlit."""
    ranges = build_ranges(df_cap)
    capacity_used = {r["RANGO_ID"]: 0.0 for r in ranges}

    data = df_data.copy()
    data["BLOQUE"] = data["PRIORIDAD"].apply(prioridad_bloque)
    data["LBS_RESTANTES"] = data["TOTAL"].astype(float)
    data["LBS_SCRAP"] = 0.0

    detalle = []
    resumen = []
    lote_id_global = 1

    block_order = ["VENCIDOS", "AHEAD", "AHEAD2", "OTROS"]

    group_keys = ["TELA.CUERPO", "MIX"]
    if "TONO" in data.columns:
        group_keys.insert(1, "TONO")
    else:
        group_keys.insert(1, "COLOR")

    groups = list(data.groupby(group_keys).groups.items())
    n_groups = max(1, len(groups))

    for gi, (keys, grp_idx) in enumerate(groups):
        if progress_cb:
            progress_cb(gi / n_groups, f"Procesando grupo {gi+1}/{n_groups}")

        work = data.loc[grp_idx].copy()
        if "TONO" in data.columns:
            tela, tono, mixv = keys[0], keys[1], keys[2]
            color = None
        else:
            tela, color, mixv = keys[0], keys[1], keys[2]
            tono = None

        ranges_mix = [r for r in ranges if r["MIX"] == mixv]
        blocked = set()

        while True:
            work["LBS_RESTANTES"] = pd.to_numeric(work["LBS_RESTANTES"], errors="coerce").fillna(0.0)
            if (work["LBS_RESTANTES"] > 0).sum() == 0:
                break
            made_any = False

            for b in block_order:
                if b in blocked:
                    continue
                cand = work[(work["BLOQUE"] == b) & (work["LBS_RESTANTES"] > 0)]
                if len(cand) == 0:
                    blocked.add(b)
                    continue

                beam_w = int(params.get("BEAM_WIDTH", 3))
                top_seeds = cand.sort_values("LBS_RESTANTES", ascending=False).head(beam_w).index.tolist()

                best_lote = None
                best_pack = None
                best_score = -1e30

                for seed_idx in top_seeds:
                    ranges_try, rule_info = reorder_ranges_for_seed(ranges_mix, mixv, work, seed_idx, params)

                    if rule_info.get("regla_aplicada") == "ANCHO18" and up(mixv) == "DYE":
                        allowed = set(params.get("ANCHO18_ALLOWED_MAX_DYE", {2200.0, 1100.0}))
                        if int(params.get("ANCHO18_ALLOW_SPILLOVER_2600", 0)) == 1:
                            allowed.add(2600.0)
                        ranges_try = [r for r in ranges_try if float(r["MAXIMO"]) in allowed]

                    lote = None
                    prioridad_obj = None

                    order_text = norm_str(params.get("WIDTHS_TARGET_ORDER", "2>3>4"))
                    targets = [int(x) for x in order_text.split(">") if x.strip().isdigit()]
                    req_strict = int(params.get("REQUIRE_WIDTHS_STRICT", 1)) == 1

                    pri_list = order_priorities(rule_info.get("prioridades", []), params)
                    use_upgrades = (len(pri_list) > 0 and int(params.get("UPGRADE_CATEGORIA", 0)) == 1)
                    pri_iter = pri_list if (use_upgrades and int(params.get("TRY_ALL_PRIORITIES", 1)) == 1) else [None]

                    for target in targets:
                        candidate_ranges_all = filter_ranges_for_width_target(ranges_try, mixv, target, params)
                        found = False
                        for pri in pri_iter:
                            candidate_ranges = candidate_ranges_all
                            if pri is not None:
                                candidate_ranges = ranges_matching_priority(pri, candidate_ranges_all, allow_nearest_higher=True)

                            for r in candidate_ranges:
                                if capacity_used[r["RANGO_ID"]] >= r["CAPACIDAD"] - 1e-6:
                                    continue
                                split_min = params.get("SPLIT_MIN_LBS_ANCHO18", 250) if rule_info.get("regla_aplicada") == "ANCHO18" else float(params.get("SPLIT_MIN_LBS_DEFAULT", 500.0))
                                intento = intentar_lote_para_rango(
                                    work, seed_idx, r, capacity_used, params, rule_info,
                                    require_two_widths=(rule_info.get("regla_aplicada") == "COMBO_ANCHOS"),
                                    split_min_lbs=split_min,
                                    min_unique_widths=target,
                                    max_unique_widths=(target if req_strict else None)
                                )
                                if intento is not None:
                                    lote = intento
                                    prioridad_obj = float(pri) if pri is not None else None
                                    found = True
                                    break
                            if found:
                                break
                        if lote is not None:
                            break

                    if lote is None:
                        if use_upgrades:
                            for pri in pri_iter:
                                if pri is None:
                                    continue
                                candidate_ranges = ranges_matching_priority(pri, ranges_try, allow_nearest_higher=True)
                                for r in candidate_ranges:
                                    if capacity_used[r["RANGO_ID"]] >= r["CAPACIDAD"] - 1e-6:
                                        continue
                                    split_min = params.get("SPLIT_MIN_LBS_ANCHO18", 250) if rule_info.get("regla_aplicada") == "ANCHO18" else float(params.get("SPLIT_MIN_LBS_DEFAULT", 500.0))
                                    if rule_info.get("regla_aplicada") == "COMBO_ANCHOS":
                                        intento = intentar_lote_para_rango(work, seed_idx, r, capacity_used, params, rule_info, require_two_widths=True, split_min_lbs=split_min)
                                        if intento is None:
                                            intento = intentar_lote_para_rango(work, seed_idx, r, capacity_used, params, rule_info, require_two_widths=False, split_min_lbs=split_min)
                                    else:
                                        intento = intentar_lote_para_rango(work, seed_idx, r, capacity_used, params, rule_info, require_two_widths=False, split_min_lbs=split_min)
                                    if intento is not None:
                                        lote = intento
                                        prioridad_obj = float(pri)
                                        break
                                if lote is not None:
                                    break

                        if lote is None:
                            for r in ranges_try:
                                if capacity_used[r["RANGO_ID"]] >= r["CAPACIDAD"] - 1e-6:
                                    continue
                                split_min = params.get("SPLIT_MIN_LBS_ANCHO18", 250) if rule_info.get("regla_aplicada") == "ANCHO18" else float(params.get("SPLIT_MIN_LBS_DEFAULT", 500.0))
                                if rule_info.get("regla_aplicada") == "COMBO_ANCHOS":
                                    intento = intentar_lote_para_rango(work, seed_idx, r, capacity_used, params, rule_info, require_two_widths=True, split_min_lbs=split_min)
                                    if intento is None:
                                        intento = intentar_lote_para_rango(work, seed_idx, r, capacity_used, params, rule_info, require_two_widths=False, split_min_lbs=split_min)
                                else:
                                    intento = intentar_lote_para_rango(work, seed_idx, r, capacity_used, params, rule_info, require_two_widths=False, split_min_lbs=split_min)
                                if intento is not None:
                                    lote = intento
                                    break

                    if lote is not None:
                        resumen_rows = []
                        for idx, _lbs, *_ in lote["ROWS"]:
                            resumen_rows.append({
                                "LNK": work.at[idx, "LNK"],
                                "ANCHOS_ROW": get_row_widths(work, idx),
                            })
                        lote_for_score = {
                            "MAXIMO": float(lote["MAXIMO"]),
                            "TOTAL_LOTE": float(lote["TOTAL_LOTE"]),
                        }
                        seed_row_dict = work.loc[seed_idx].to_dict()
                        sc = score_lote(lote_for_score, resumen_rows, params, categoria=lote["CATEGORIA"], seed_row=seed_row_dict)
                        if sc > best_score:
                            best_score = sc
                            best_lote = lote
                            best_pack = (lote, rule_info, prioridad_obj, best_score)

                if best_lote is None:
                    blocked.add(b)
                    continue

                lote, rule_info, prioridad_obj, best_score = best_pack
                split_min = params.get("SPLIT_MIN_LBS_ANCHO18", 250) if rule_info.get("regla_aplicada") == "ANCHO18" else float(params.get("SPLIT_MIN_LBS_DEFAULT", 500.0))

                lote_id = f"L{lote_id_global:06d}"
                lote_id_global += 1

                lote_widths = []
                for idx, _lbs, *_ in lote["ROWS"]:
                    lote_widths += get_row_widths(work, idx)
                anchos_lote = sorted(set([float(w) for w in lote_widths if w is not None and not pd.isna(w) and float(w) != 0.0]))
                anchos_lote_str = str(anchos_lote)

                prioridad_final = float(lote["MAXIMO"])
                regla_aplicada_final = rule_info.get("regla_aplicada", "NONE")
                requiere_2_anchos_flag = False
                if regla_aplicada_final == "COMBO_ANCHOS":
                    requiere_2_anchos_flag = bool(lote.get("REQUIERE_2_ANCHOS", False)) and (len(anchos_lote) >= 2)
                    if not requiere_2_anchos_flag:
                        regla_aplicada_final = "COMBO_ANCHOS_FALLBACK"

                for idx, lbs_asig, over_extra, under_saved in lote["ROWS"]:
                    detalle.append({
                        "LOTE_ID": lote_id,
                        "ANCHOS_LOTE": anchos_lote_str,
                        "CATEGORIA": lote["CATEGORIA"],
                        "MIX": lote["MIX"],
                        "TELA.CUERPO": tela,
                        "COLOR": work.at[idx, "COLOR"],
                        "TONO": work.at[idx, "TONO"] if "TONO" in work.columns else "",
                        "LNK_PRIORIDAD": f"{work.at[idx,'LNK']}|{work.at[idx,'PRIORIDAD']}",
                        "LNK": work.at[idx, "LNK"],
                        "PRIORIDAD": work.at[idx, "PRIORIDAD"],
                        "BLOQUE": work.at[idx, "BLOQUE"],
                        "ANCHO.F.C": float(work.at[idx, "ANCHO.F.C"]),
                        "ANCHO.F.M": float(work.at[idx, "ANCHO.F.M"]),
                        "CONSUMO_C": float(work.at[idx, "CONSUMO_C"]),
                        "FAMILIA": work.at[idx, "FAMILIA"],
                        "COLOR_R": work.at[idx, "COLOR_R"],
                        "STYLE": work.at[idx, "STYLE"],
                        "TIPO_TEJIDO": work.at[idx, "TIPO_TEJIDO"] if "TIPO_TEJIDO" in work.columns else "",
                        "PCT_CARGA": work.at[idx, "PCT_CARGA"] if "PCT_CARGA" in work.columns else 1.0,
                        "LBS_ASIGNADAS": float(lbs_asig),
                        "LBS_EXTRA_SOBRE_ORDEN": float(max(0.0, over_extra)),
                        "APLICA_REGLA": regla_aplicada_final,
                        "PRIORIDAD_USADA": prioridad_final,
                        "PRIORIDAD_OBJETIVO": prioridad_obj,
                        "ORIGEN_PRIORIDAD": rule_info.get("origen_prioridad", "MIX"),
                        "MATCH_ANCHO": bool(rule_info.get("match_combo", False)),
                        "LIMITE_ANCHO_STYLE": rule_info.get("limite_ancho_style", None),
                        "UPGRADE_CATEGORIA": int(params.get("UPGRADE_CATEGORIA", 0)),
                        "SPLIT_MIN_USADO": float(split_min),
                        "REQUIERE_2_ANCHOS": bool(requiere_2_anchos_flag),
                        "DECISION_SCORE": float(best_score)
                    })

                    prev_rest = float(work.at[idx, "LBS_RESTANTES"])
                    new_rest = prev_rest - float(lbs_asig)
                    work.at[idx, "LBS_RESTANTES"] = max(0.0, new_rest)

                    if int(params.get("SCRAP_REMAINDER_BELOW_SPLIT_MIN", 1)) == 1:
                        rem = float(work.at[idx, "LBS_RESTANTES"])
                        if rem > 1e-9 and rem + 1e-9 < float(split_min):
                            work.at[idx, "LBS_SCRAP"] = float(work.at[idx, "LBS_SCRAP"]) + rem
                            work.at[idx, "LBS_RESTANTES"] = 0.0

                det_lote = [d for d in detalle if d["LOTE_ID"] == lote_id]
                lnks = {d["LNK"] for d in det_lote}
                bloques = [d["BLOQUE"] for d in det_lote]
                bloque_dom = max(set(bloques), key=bloques.count) if bloques else ""

                resumen.append({
                    "LOTE_ID": lote_id,
                    "ANCHOS_LOTE": anchos_lote_str,
                    "CATEGORIA": lote["CATEGORIA"],
                    "MIX": lote["MIX"],
                    "TELA.CUERPO": tela,
                    "COLOR/TONO_KEY": tono if tono is not None else color,
                    "LBS_TOTAL": float(lote["TOTAL_LOTE"]),
                    "MIN_RANGO": float(lote["MINIMO"]),
                    "MAX_RANGO": float(lote["MAXIMO"]),
                    "CAPACIDAD_PERDIDA": float(lote["MAXIMO"] - lote["TOTAL_LOTE"]),
                    "SKU_DISTINTOS": len(lnks),
                    "ANCHOS_UNICOS": len(anchos_lote),
                    "BLOQUE_DOMINANTE": bloque_dom,
                    "REGLA_DOMINANTE": regla_aplicada_final,
                    "PRIORIDAD_FINAL": prioridad_final,
                    "PRIORIDAD_OBJETIVO": prioridad_obj,
                    "COMBO_ANCHOS": (regla_aplicada_final == "COMBO_ANCHOS"),
                    "STYLE_CRITICO": True if rule_info.get("regla_aplicada") == "ANCHO18" else False,
                    "CANT_REGLAS_APLICADAS": 0 if rule_info.get("regla_aplicada") == "NONE" else 1,
                    "UPGRADE_CATEGORIA": int(params.get("UPGRADE_CATEGORIA", 0)),
                    "PCT_CARGA_USADO": lote.get("PCT_CARGA_USADO", 1.0),
                })

                capacity_used[lote["RANGO_ID"]] += float(lote["TOTAL_LOTE"])
                blocked = set()
                made_any = True
                break

            if not made_any:
                break

        data.loc[work.index, "LBS_RESTANTES"] = work["LBS_RESTANTES"]
        data.loc[work.index, "LBS_SCRAP"] = work["LBS_SCRAP"]

    if progress_cb:
        progress_cb(1.0, "Loteo finalizado")

    exced_cols = ["LNK", "TELA.CUERPO", "COLOR", "MIX", "PRIORIDAD", "BLOQUE", "ANCHO.F.C", "ANCHO.F.M", "TOTAL", "LBS_RESTANTES", "LBS_SCRAP"]
    if "TONO" in data.columns:
        exced_cols.insert(3, "TONO")
    exced = data[data["LBS_RESTANTES"] > 1e-9][exced_cols].copy()

    df_detalle = pd.DataFrame(detalle)
    if len(df_detalle) > 0:
        df_detalle["DOCENAS"] = np.where(df_detalle["CONSUMO_C"] > 0, df_detalle["LBS_ASIGNADAS"] / df_detalle["CONSUMO_C"], np.nan)
    df_resumen = pd.DataFrame(resumen)

    df_param_out = pd.DataFrame([
        ["MIN_DIFF", params["MIN_DIFF"]],
        ["MAX_DIFF", params["MAX_DIFF"]],
        ["MAX_WIDTHS_BY_CAT", str(params.get("MAX_WIDTHS_BY_CAT", {}))],
        ["MAX_SKU", params["MAX_SKU"]],
        ["SPLIT_MIN_LBS_DEFAULT", params.get("SPLIT_MIN_LBS_DEFAULT", 500.0)],
        ["SPLIT_MIN_LBS_ANCHO18", params.get("SPLIT_MIN_LBS_ANCHO18", 250)],
        ["RULE_ORDER", params.get("RULE_ORDER", "")],
        ["PRIORITY_ORDER", params.get("PRIORITY_ORDER", "")],
        ["APPLY_RULES_BLEACH", params.get("APPLY_RULES_BLEACH", 0)],
        ["TRY_ALL_PRIORITIES", params.get("TRY_ALL_PRIORITIES", 1)],
        ["UPGRADE_CATEGORIA", params.get("UPGRADE_CATEGORIA", 0)],
        ["ANCHO18_ALLOW_SPILLOVER_2600", params.get("ANCHO18_ALLOW_SPILLOVER_2600", 0)],
        ["ANCHO18_ALLOWED_MAX_DYE", ",".join(sorted(str(int(x)) for x in params.get("ANCHO18_ALLOWED_MAX_DYE", {2200.0, 1100.0})))],
        ["SCRAP_REMAINDER_BELOW_SPLIT_MIN", params.get("SCRAP_REMAINDER_BELOW_SPLIT_MIN", 1)],
        ["BEAM_WIDTH", params.get("BEAM_WIDTH", 3)],
        ["W_FILL", params.get("W_FILL", 5.0)],
        ["W_CAP_LOSS", params.get("W_CAP_LOSS", 3.0)],
        ["WIDTH_PREF_LIST", ",".join(str(x) for x in params.get("WIDTH_PREF_LIST", [2, 3, 1, 4, 5, 6]))],
        ["W_WIDTH_PREF", params.get("W_WIDTH_PREF", 2.0)],
        ["W_1100_WIDTHS_STRICT", params.get("W_1100_WIDTHS_STRICT", 10.0)],
        ["WIDTHS_TARGET_ORDER", params.get("WIDTHS_TARGET_ORDER", "2>3>4")],
        ["REQUIRE_WIDTHS_STRICT", params.get("REQUIRE_WIDTHS_STRICT", 1)],
        ["ALLOWED_MAXIMO_FOR_3_WIDTHS_DYE", ",".join(str(int(x)) for x in params.get("ALLOWED_MAXIMO_FOR_3_WIDTHS", {}).get("DYE", set()))],
        ["ALLOWED_MAXIMO_FOR_4_WIDTHS_DYE", ",".join(str(int(x)) for x in params.get("ALLOWED_MAXIMO_FOR_4_WIDTHS", {}).get("DYE", set()))],
        ["ALLOWED_MAXIMO_FOR_3_WIDTHS_BLEACH", ",".join(str(int(x)) for x in params.get("ALLOWED_MAXIMO_FOR_3_WIDTHS", {}).get("BLEACH", set()))],
        ["ALLOWED_MAXIMO_FOR_4_WIDTHS_BLEACH", ",".join(str(int(x)) for x in params.get("ALLOWED_MAXIMO_FOR_4_WIDTHS", {}).get("BLEACH", set()))],
        ["TIPO_TEJIDO_ENABLE", params.get("TIPO_TEJIDO_ENABLE", 0)],
        ["W_TIPO_TEJIDO_FLEECE", params.get("W_TIPO_TEJIDO_FLEECE", 4.0)],
    ], columns=["PARAMETRO", "VALOR"])

    return df_detalle, df_resumen, exced, df_param_out

# ---------------------------- Reportes adicionales ----------------------------
def build_reports(df_data, df_cap, df_detalle, df_resumen):
    df_cap_simple = df_cap[["CATEGORIA", "MIX", "MINIMO", "MAXIMO", "CAPACIDAD"]].copy()
    if len(df_detalle) > 0:
        df_cat_asig = df_detalle.groupby(["CATEGORIA", "MIX"], as_index=False)["LBS_ASIGNADAS"].sum()
    else:
        df_cat_asig = pd.DataFrame({"CATEGORIA": [], "MIX": [], "LBS_ASIGNADAS": []})
    df_cap_cap = (df_cap_simple.merge(df_cat_asig, on=["CATEGORIA", "MIX"], how="left")
                  .fillna({"LBS_ASIGNADAS": 0.0}))
    df_cap_cap["DIFERENCIA"] = df_cap_cap["LBS_ASIGNADAS"] - df_cap_cap["CAPACIDAD"]
    df_cap_cap["FILL_RATE"] = np.where(df_cap_cap["CAPACIDAD"] > 0, df_cap_cap["LBS_ASIGNADAS"] / df_cap_cap["CAPACIDAD"], 0.0)
    df_cap_cap = df_cap_cap.sort_values(["MIX", "CATEGORIA"])

    df_base_blocks = df_data.copy()
    df_base_blocks["BLOQUE"] = df_base_blocks["PRIORIDAD"].apply(prioridad_bloque)
    df_prio_base = (df_base_blocks.groupby(["MIX", "BLOQUE"], as_index=False)["TOTAL"].sum()
                     .rename(columns={"TOTAL": "LBS_BASE"}))

    if len(df_detalle) > 0:
        df_prio_asig = df_detalle.groupby(["MIX", "BLOQUE"], as_index=False)["LBS_ASIGNADAS"].sum()
    else:
        df_prio_asig = pd.DataFrame({"MIX": [], "BLOQUE": [], "LBS_ASIGNADAS": []})

    df_prio_vs_asig = (df_prio_base.merge(df_prio_asig, on=["MIX", "BLOQUE"], how="left")
                        .fillna({"LBS_ASIGNADAS": 0.0}))
    df_prio_vs_asig["LBS_SIN_ASIGNAR"] = df_prio_vs_asig["LBS_BASE"] - df_prio_vs_asig["LBS_ASIGNADAS"]
    order_blocks = ["VENCIDOS", "AHEAD", "AHEAD2", "OTROS"]
    df_prio_vs_asig["ORD"] = df_prio_vs_asig["BLOQUE"].apply(lambda x: order_blocks.index(x) if x in order_blocks else 999)
    df_prio_vs_asig = df_prio_vs_asig.sort_values(["MIX", "ORD"]).drop(columns=["ORD"])

    df_lnk_base = (df_data.groupby(["MIX", "LNK"], as_index=False)["TOTAL"].sum()
                    .rename(columns={"TOTAL": "LBS_BASE"}))
    if "LBS_SCRAP" in df_data.columns:
        df_lnk_scrap = (df_data.groupby(["MIX", "LNK"], as_index=False)["LBS_SCRAP"].sum())
    else:
        df_lnk_scrap = pd.DataFrame({"MIX": [], "LNK": [], "LBS_SCRAP": []})

    if len(df_detalle) > 0:
        df_lnk_asig = df_detalle.groupby(["MIX", "LNK"], as_index=False)["LBS_ASIGNADAS"].sum()
    else:
        df_lnk_asig = pd.DataFrame({"MIX": [], "LNK": [], "LBS_ASIGNADAS": []})

    df_lnk_comp = (df_lnk_base.merge(df_lnk_asig, on=["MIX", "LNK"], how="left")
                   .merge(df_lnk_scrap, on=["MIX", "LNK"], how="left")
                   .fillna({"LBS_ASIGNADAS": 0.0, "LBS_SCRAP": 0.0}))
    df_lnk_comp["BALANCE"] = df_lnk_comp["LBS_BASE"] - df_lnk_comp["LBS_ASIGNADAS"] - df_lnk_comp["LBS_SCRAP"]
    df_lnk_comp["ESTADO"] = np.where(df_lnk_comp["BALANCE"].abs() <= 1e-6,
                                      np.where(df_lnk_comp["LBS_SCRAP"] > 1e-6, "COMPLETO (SCRAP)", "COMPLETO"),
                                      "INCOMPLETO")
    df_lnk_comp = df_lnk_comp.sort_values(["MIX", "ESTADO", "BALANCE"], ascending=[True, True, False])

    def resumen_por_lote(df_det, filtro):
        if len(df_det) == 0:
            return pd.DataFrame()
        sub = df_det.query(filtro).copy()
        if len(sub) == 0:
            return pd.DataFrame()
        agg = (sub.groupby("LOTE_ID", as_index=False)
               .agg({
                   "ANCHOS_LOTE": "first",
                   "MIX": "first",
                   "TELA.CUERPO": "first",
                   "COLOR": "first",
                   "TONO": "first" if "TONO" in df_det.columns else (lambda x: ""),
                   "FAMILIA": "first",
                   "STYLE": "first",
                   "COLOR_R": "first",
                   "PRIORIDAD_USADA": "first",
                   "PRIORIDAD_OBJETIVO": "first",
                   "UPGRADE_CATEGORIA": "first",
                   "LBS_ASIGNADAS": "sum"
               }))
        return agg

    rep_ancho18 = resumen_por_lote(df_detalle, "APLICA_REGLA == 'ANCHO18'")
    rep_combo = resumen_por_lote(df_detalle, "APLICA_REGLA == 'COMBO_ANCHOS'")
    rep_color = resumen_por_lote(df_detalle, "APLICA_REGLA == 'COLOR_R'")
    rep_fam = resumen_por_lote(df_detalle, "APLICA_REGLA == 'FAMILIA'")

    if len(df_resumen) > 0:
        rep_maestro = df_resumen[[
            "LOTE_ID", "ANCHOS_LOTE", "MIX", "REGLA_DOMINANTE", "PRIORIDAD_FINAL", "PRIORIDAD_OBJETIVO",
            "COMBO_ANCHOS", "STYLE_CRITICO", "CANT_REGLAS_APLICADAS",
            "ANCHOS_UNICOS", "LBS_TOTAL", "CAPACIDAD_PERDIDA", "UPGRADE_CATEGORIA"
        ]].copy()
    else:
        rep_maestro = pd.DataFrame()

    if len(df_detalle) > 0:
        overs = (df_detalle.groupby(["MIX", "LNK"], as_index=False)
                 .agg({"LBS_EXTRA_SOBRE_ORDEN": "sum", "LBS_ASIGNADAS": "sum"}))
        tono_col = "TONO" if "TONO" in df_detalle.columns else "LNK"
        decision_log = df_detalle[[
            "LOTE_ID", "MIX", "TELA.CUERPO", "COLOR", tono_col,
            "FAMILIA", "STYLE", "COLOR_R",
            "CATEGORIA", "PRIORIDAD", "BLOQUE", "APLICA_REGLA", "PRIORIDAD_USADA", "PRIORIDAD_OBJETIVO",
            "LBS_ASIGNADAS", "LBS_EXTRA_SOBRE_ORDEN", "ANCHOS_LOTE", "DECISION_SCORE", "LNK"
        ]].copy().sort_values(["LOTE_ID", "LNK"])
    else:
        overs = pd.DataFrame({"MIX": [], "LNK": [], "LBS_EXTRA_SOBRE_ORDEN": [], "LBS_ASIGNADAS": []})
        decision_log = pd.DataFrame()

    return {
        "CAPACIDAD_X_CATEG": df_cap_cap,
        "PRIORIDAD_VS_ASIG": df_prio_vs_asig,
        "LNK_COMPLETITUD": df_lnk_comp,
        "REGLA_STYLE_ANCHO18": rep_ancho18,
        "REGLA_COMBINACION_ANCHOS": rep_combo,
        "REGLA_COLOR_R": rep_color,
        "REGLA_FAMILIA": rep_fam,
        "REPORTE_REGLAS_MIX": rep_maestro,
        "OVERSHOOT_SUMMARY": overs,
        "DECISION_LOG": decision_log
    }

# ---------------------------- Formatting / export ----------------------------
def format_workbook(path_xlsx, font_name="Cambria", font_size=8):
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path_xlsx)
    f = Font(name=font_name, size=font_size)
    for ws in wb.worksheets:
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).font = f
        for c in range(1, ws.max_column + 1):
            col = get_column_letter(c)
            max_len = 0
            for r in range(1, ws.max_row + 1):
                v = ws.cell(r, c).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col].width = min(max(8, max_len + 2), 60)
    wb.save(path_xlsx)
